"""
╔════════════════════════════════════════════════════════════════════════════╗
║                  ПРОГРАММА 2: СНС «ОПРЕДЕЛЕНИЕ ТРАЕКТОРИЙ                 ║
║                   ДВИЖЕНИЯ ПАССАЖИРОВ В ТРАНСПОРТЕ»                       ║
╚════════════════════════════════════════════════════════════════════════════╝

ПОЛНОЕ ОПИСАНИЕ СИСТЕМЫ:

СНС реализует интеллектуальное отслеживание пассажиров в общественном
транспорте (автобус, троллейбус, трамвай) с использованием видеокамеры.

ОСНОВНЫЕ МОДУЛИ И ФУНКЦИИ:

1. ДЕТЕКЦИЯ И ТРЕКИНГ ОБЪЕКТОВ (YOLO + трекинг центра)
   • Запуск свёрточных нейросетей YOLOv8n для обнаружения людей на видео
   • Оптимизация: запуск YOLO через N кадров, остальные используют кэш
   • Трекинг центра объекта между кадрами для выявления пересечения линии

2. ИДЕНТИФИКАЦИЯ ПАССАЖИРОВ (День 2–7)
   • Извлечение параметров лица: ориентация (YPR), контур, расстояния
   • 214-мерный дескриптор лица на основе MediaPipe FaceMesh
   • Multi-modal matching: 70% вес лицо + 30% вес одежда
   • Анализ силуэта: рост, цвета одежды (RGB), тип одежды
   • Пороги калибруются в реальном времени горячими клавишами

3. OD-МАТРИЦА КОРРЕСПОНДЕНЦИЙ (День 4–5)
   • Регистрация входа/выхода пассажира на каждой остановке
   • Матрица привязана к остановкам и направлениям движения
   • Сбор временных данных для анализа продолжительности поездок
   • Экспорт в JSON/CSV/Excel с цветным оформлением

4. ТРАЕКТОРИИ И ЦЕПОЧКИ МАРШРУТОВ (День 3, День 10)
   • Полный путь каждого пассажира по остановкам
   • Ночная сшивка: объединение поездок одного человека через маршруты
   • Cross-route stitching: поиск пересадок (время и остановка совпадают)
   • Вывод: единые цепочки пассажира с начальной и конечной точками

5. АНАЛИЗ ТРАНЗИТНЫХ ПАССАЖИРОВ (День 11+)
   • Новое: выявление пассажиров, вошедших и вышедших на ОДНОЙ остановке
   • Если время на борту ≤ TIME_AT_STOP_SEC → считается внутристанционным
     переходом/ошибкой обнаружения, отдельная статистика в отчётах
   • Настраивается параметром TIME_AT_STOP_SEC (сек)

6. ВЕБ-ДАШБОРД В РЕАЛЬНОМ ВРЕМЕНИ (День 6)
   • HTTP-сервер на порту 8888
   • Live-данные: счётчики, OD-матрица, параметры пассажиров
   • HTML-таблицы с цветным оформлением

7. ОТЧЁТНОСТЬ И ЭКСПОРТ (День 5, День 10)
   • CSV: по дням, часам, маршрутам, остановкам
   • Excel: единый файл с 12+ листами, форматирование, графики
   • JSON: для автоматической обработки

8. СТАБИЛЬНОСТЬ И МОНИТОРИНГ (День 8, 11)
   • FPS-оптимизация: YOLO_EVERY_N_FRAMES, кэш боксов
   • Latency-мониторинг: вывод задержек YOLO и полного цикла
   • Горячие клавиши для калибровки порогов в real-time
   • Watchdog-поток для отслеживания зависаний
   • Graceful shutdown: SIGINT/SIGTERM обработчики, сохранение конфига
   • Auto-reconnect для сетевых потоков (RTSP/HTTP)

АРХИТЕКТУРА:
  - Основной цикл: захват видео → обработка YOLO → трекинг → идентификация
  - Параллельные потоки: веб-сервер, анализ видео (ThreadedVideoReader)
  - БД (SQLite): passengers, trajectories, od_log, passenger_flow
  - Потокобезопасность: threading.Lock для доступа к счётчикам

УПРАВЛЕНИЕ:
  q/й       — выход
  h/р       — горизонтальная линия
  v/м       — вертикальная линия
  c/с       — сброс счётчиков и перерисовка линии
  PgUp/,    — предыдущая остановка
  PgDn/.    — следующая остановка
  s/ы       — управление остановками (выпадающий список)
  r/к       — генерировать отчёты
  x/ч       — запустить ночную сшивку маршрутов
  
  [ / ]     — YOLO_CONFIDENCE ±0.05
  - / =     — PASSENGER_ID_THRESHOLD ±0.02
  ; / '     — TIME_AT_STOP_SEC ±5 сек (время стоянки)
  p/з       — профиль производительности (ОЗУ, FPS, треки)
  b/и       — создать файлы сборки EXE (PyInstaller)
  m/ь       — сгенерировать инструкцию пользователя

9.  ПРОФИЛИРОВАНИЕ И МОНИТОРИНГ (День 12)
   • PerformanceProfiler: ОЗУ, FPS мин/макс/сред, пик треков
   • Замер через Windows API / psutil (fallback)
   • Периодическая очистка устаревших треков (TRACK_CLEANUP_FRAMES)
   • Вывод отчёта по клавише P

10. СБОРКА В EXE (День 13)
   • Генерация People_Counter_v2.spec для PyInstaller
   • batch-скрипт build_exe_v2.bat: CPU-only torch, UPX-сжатие
   • Все модули и модель пакуются в один каталог

11. ДОКУМЕНТАЦИЯ ПОЛЬЗОВАТЕЛЯ (День 14)
   • Генерация ИНСТРУКЦИЯ_ПРОГ2.txt по клавише M
   • 9 разделов: назначение, установка, управление, отчёты, FAQ

12. САМОДИАГНОСТИКА И ФИНАЛИЗАЦИЯ (День 15)
   • run_self_diagnostic(): проверка всех компонентов перед запуском
   • print_session_summary(): отчёт сессии с рекомендациями
   • Таблица готовности: модули, БД, модель, шрифты

ФОРМАТЫ ДАННЫХ:
  • Дескриптор лица: JSON-массив 214 float32
  • Дескриптор внешности: JSON-массив 10 float32 (рост, RGB, цвета)
  • Времена: ISO 8601 (%Y-%m-%d %H:%M:%S)
"""

import cv2
import numpy as np
import os
import signal
import sqlite3
import sys
import threading
import json
from datetime import datetime
from http.server import HTTPServer, BaseHTTPRequestHandler
import time
from urllib.parse import urlparse

try:
    import pandas as pd
except ImportError:
    pd = None

try:
    from PIL import Image, ImageDraw, ImageFont
except ImportError:
    Image = None
    ImageDraw = None
    ImageFont = None

try:
    from ultralytics import YOLO
except ImportError:
    YOLO = None

try:
    from passenger_identifier import PassengerIdentifier, PassengerDB, ProfileAnalyzer
    IDENT_AVAILABLE = True
except ImportError:
    IDENT_AVAILABLE = False

try:
    from od_matrix_tracker import ODMatrixTracker
    OD_AVAILABLE = True
except ImportError:
    OD_AVAILABLE = False

APP_DIR = os.path.dirname(os.path.abspath(sys.executable if getattr(sys, 'frozen', False) else __file__))


def app_path(*parts):
    return os.path.join(APP_DIR, *parts)

# ─── Источник видео ───────────────────────────────────────────────────────────
# Варианты:
#   0                          — встроенная/USB-камера
#   'video.mp4'                — видеофайл
#   PHONE_STREAM               — телефон по Wi-Fi (см. настройки ниже)
#   'rtsp://user:pass@IP:port' — IP-камера по RTSP

# --- Подключение телефона (Android: IP Webcam, iPhone: IP Camera Lite) --------
# 1. Установите на телефон приложение:
#      Android: «IP Webcam» (Pavel Khlebovich) — бесплатно
#      iPhone:  «IP Camera Lite»               — бесплатно
# 2. Запустите сервер в приложении, запишите IP и порт (обычно 8080)
# 3. Вставьте IP телефона ниже и выберите нужный формат потока:

PHONE_IP   = '192.168.0.100'   # ← замените на IP вашего телефона
PHONE_PORT = 8080

# Форматы (раскомментируйте нужный):
PHONE_STREAM = f'http://{PHONE_IP}:{PHONE_PORT}/video'          # MJPEG — проще, работает везде
# PHONE_STREAM = f'rtsp://{PHONE_IP}:{PHONE_PORT}/h264_ulaw.sdp'  # RTSP H.264 — меньше задержка
PHONE_FALLBACK_PATHS = ['/video', '/videofeed', '/video/mjpeg', '/mjpeg']

VIDEO_SOURCE = PHONE_STREAM   # ← переключите на 0 для обычной камеры
AUTO_START_LAST_SOURCE = True
FORCE_SOURCE_DIALOG = any(arg.lower() in ('--source', '--choose-source') for arg in sys.argv[1:])

# ─── Организация ─────────────────────────────────────────────────────────────
ORG_NAME  = 'НМУ ВКС'
ORG_PHONE = '+7 952 553-96-21'
ORG_EMAIL = 'mveo@yandex.ru'
ORG_CITY  = 'Воронеж'

# ─── Маршрут и транспорт ─────────────────────────────────────────────────────
ROUTE_NAME   = 'Маршрут №1'
VEHICLE_NAME = 'Автобус 001'
STOP_NAME    = 'Остановка А'
DOOR_NUMBER  = 1

# Список остановок маршрута
STOP_LIST = ['Остановка А', 'Остановка Б', 'Остановка В']
current_stop_index = 0

# Счётчики по каждой остановке: {"Остановка А": {"enter": 0, "exit": 0}, ...}
stop_counters = {s: {'enter': 0, 'exit': 0} for s in STOP_LIST}


def rebuild_stop_counters():
    global stop_counters
    stop_counters = {s: stop_counters.get(s, {'enter': 0, 'exit': 0}) for s in STOP_LIST}


def parse_stop_list(raw_text):
    text = str(raw_text or '').replace('\r', '\n').replace(',', ';').replace('\n', ';')
    stops = [item.strip() for item in text.split(';') if item.strip()]
    return stops or ['Остановка А']


def apply_transport_settings(route_name=None, vehicle_name=None, door_number=None,
                             stop_list=None, start_stop_index=None):
    global ROUTE_NAME, VEHICLE_NAME, DOOR_NUMBER
    global STOP_LIST, current_stop_index, STOP_NAME

    if route_name is not None:
        ROUTE_NAME = str(route_name).strip() or ROUTE_NAME
    if vehicle_name is not None:
        VEHICLE_NAME = str(vehicle_name).strip() or VEHICLE_NAME
    if door_number is not None:
        try:
            DOOR_NUMBER = max(1, int(door_number))
        except Exception:
            pass
    if stop_list is not None:
        STOP_LIST = parse_stop_list(stop_list)
        rebuild_stop_counters()
    if start_stop_index is not None:
        try:
            current_stop_index = max(0, min(int(start_stop_index), len(STOP_LIST) - 1))
        except Exception:
            current_stop_index = 0
    else:
        current_stop_index = max(0, min(current_stop_index, len(STOP_LIST) - 1))
    STOP_NAME = STOP_LIST[current_stop_index]
    load_stop_counters(STOP_NAME)


def save_current_stop_counters():
    """Save enter/exit counts to current stop before switching."""
    global enter_count, exit_count
    stop_counters[STOP_NAME] = {'enter': enter_count, 'exit': exit_count}


def load_stop_counters(stop_name):
    """Load enter/exit counts from a stop."""
    global enter_count, exit_count
    if stop_name not in stop_counters:
        stop_counters[stop_name] = {'enter': 0, 'exit': 0}
    enter_count = stop_counters[stop_name]['enter']
    exit_count = stop_counters[stop_name]['exit']


def switch_to_stop(idx):
    """Switch to stop by index, saving current and loading new counters."""
    global current_stop_index, STOP_NAME
    if 0 <= idx < len(STOP_LIST) and idx != current_stop_index:
        save_current_stop_counters()
        current_stop_index = idx
        STOP_NAME = STOP_LIST[current_stop_index]
        load_stop_counters(STOP_NAME)


def get_total_counters():
    """Return totals across all stops."""
    save_current_stop_counters()  # ensure current is saved
    total_in = sum(c['enter'] for c in stop_counters.values())
    total_out = sum(c['exit'] for c in stop_counters.values())
    return total_in, total_out


def get_stop_stats():
    """Return list of per-stop stats for API/dashboard."""
    save_current_stop_counters()
    stats = []
    for s in STOP_LIST:
        c = stop_counters.get(s, {'enter': 0, 'exit': 0})
        stats.append({'name': s, 'enter': c['enter'], 'exit': c['exit']})
    return stats

# ─── Линия ───────────────────────────────────────────────────────────────────
LINE_ORIENTATION = 'vertical'  # 'horizontal' или 'vertical'

# Горизонтальная линия
LINE_Y = 300
LINE_START_X = 100
LINE_END_X = 900

# Вертикальная линия
LINE_X = 500
LINE_START_Y = 100
LINE_END_Y = 700

MIN_AREA = 700  # Для удалённых/узких ракурсов (лестница) ниже порог лучше захватывает людей
YOLO_CONFIDENCE = 0.28  # Чуть мягче порог для реальных условий камеры
YOLO_MODEL = app_path('yolov8n.onnx')  # ONNX-формат: работает без AVX2, совместим со старыми CPU
LINE_REARM_DISTANCE = 24  # быстрее повторное срабатывание после ухода от линии
LINE_KEYBOARD_STEP = 18
LINE_CLICK_MOVE_THRESHOLD = 8
COUNT_POINT_MODE = 'bottom'  # 'center' или 'bottom': нижняя точка устойчивее к маханию руками
REVERSE_COUNT_DIRECTION = False  # False=как есть, True=вход/выход поменять местами
CAMERA_ROTATION = 0  # 0/90/180/270 программный разворот кадра

# ─── Настройки отображения ────────────────────────────────────────────────────
# Любая камера/видео масштабируется к этому размеру до обработки.
# Это гарантирует, что окно никогда не выйдет за пределы монитора,
# а линия и трекинг работают одинаково при любом исходном разрешении.
DISPLAY_W = 900
DISPLAY_H = 500
PANEL_H = 150
# Автоподгонка финального окна (кадр + панель) под экран.
AUTO_FIT_TO_SCREEN = True
WINDOW_MARGIN_W = 120
WINDOW_MARGIN_H = 120

# ─── Трекинг ─────────────────────────────────────────────────────────────────
# Максимальное смещение центра объекта между кадрами (доля ширины экрана).
TRACKING_MAX_DIST_RATIO = 0.16   # больше допуск для лестницы и неровного шага
# Сколько кадров сохранять «потерянный» трек (защита от пропуска YOLO).
TRACKING_LOST_FRAMES = 14

# ─── День 8: Оптимизация производительности ──────────────────────────────────
# YOLO запускается не каждый кадр, а каждые N-й.
# Остальные кадры используют боксы с предыдущего YOLO-запуска.
# Это даёт прирост FPS x2–x3 без заметной потери точности.
YOLO_EVERY_N_FRAMES = 1          # 1 = каждый кадр (макс. точность),
                                  # 2–3 = баланс точности и скорости
YOLO_IMGSZ = 416                  # размер входа YOLO (ниже = быстрее)
YOLO_HALF = False                 # FP16 (True = быстрее на GPU, False = CPU-safe)

# ─── Настройки захвата для USB/web-камеры ────────────────────────────────────
# SAFE_MODE=True не форсирует MJPG — убирает мигание/полосы на проблемных драйверах.
USB_CAM_SAFE_MODE = True
USB_CAM_TARGET_FPS = 20
SHOW_SPLASH_SCREEN = False
ENABLE_WEB_DASHBOARD = False
OD_TRACKING_ENABLED = False

# ─── Идентификация пассажиров ────────────────────────────────────────────────
PASSENGER_ID_ENABLED = False        # на слабом ПК отключено: нужен стабильный счёт, а не биометрия
PASSENGER_ID_THRESHOLD = 0.22       # порог схожести (меньше = строже, 0.15–0.30)
IDENTIFY_EVERY_N_FRAMES = 10        # обновлять дескриптор трека каждые N кадров

# ─── День 9: Калибровка порогов ──────────────────────────────────────────────
# Порог YOLO-уверенности и идентификации подстраиваются горячими клавишами:
#   [ / ] — уменьшить/увеличить YOLO_CONFIDENCE на 0.05
#   - / = — уменьшить/увеличить PASSENGER_ID_THRESHOLD на 0.02
# Текущие значения порогов отображаются на панели.

# ─── Анализ профиля/силуэта (рост, одежда, цвета) ───────────────────
PROFILE_ANALYSIS_ENABLED = False     # на слабом ПК отключено ради плавности интерфейса
CAMERA_HEIGHT_M = 2.5                # высота установки камеры (м)
VISIBLE_HEIGHT_M = 2.2               # видимая высота зоны кадра (м)
APPEARANCE_EVERY_N_FRAMES = 30       # обновлять параметры внешности каждые N кадров

# Автопереподключение сетевого потока при обрыве (RTSP/HTTP).
AUTO_RECONNECT_STREAM = True
RECONNECT_NOFRAME_SEC = 2.5
RECONNECT_COOLDOWN_SEC = 3.0

# База данных и отчёты
DATABASE_PATH = app_path('passenger_flow.db')
REPORT_FOLDER = app_path('reports')

# ─── Анализ транзитных пассажиров (Пожелание заказчика) ──────────────────────
# Если пассажир вошёл и вышел на ОДНОЙ остановке за время ≤ TIME_AT_STOP_SEC (сек),
# это считается внутристанционным переходом/ошибкой обнаружения.
# Такие пассажиры выводятся в отдельной колонке в отчётах.
TIME_AT_STOP_SEC = 30.0              # порог времени на борту (сек) для выявления транзитов

draw_state = {
    'drawing': False,
    'start': None,
    'end': None
}

# Координаты кнопок на панели (обновляются каждый кадр)
panel_buttons = {}   # {'prev': (x1,y1,x2,y2), 'next': (x1,y1,x2,y2), 'manage': (x1,y1,x2,y2), 'settings': (...)}
frame_height = 0     # высота кадра (для пересчёта Y в панель)
frame_width = 0      # ширина кадра (для перемещения линии мышью)
settings_request = False  # открыть общее окно настроек из панели/горячей клавиши

# ─── День 11: Стабильность — сохранение/загрузка конфигурации ────────────────
CONFIG_PATH = app_path('people_counter_config.json')
_config_loaded = False
# Флаг штатного завершения (SIGINT / клавиша Q)
_shutdown_requested = False

def save_config():
    """День 11: сохраняет текущие пороги и параметры в JSON-файл."""
    cfg = {
        'VIDEO_SOURCE': VIDEO_SOURCE,
        'PHONE_IP': PHONE_IP,
        'PHONE_PORT': PHONE_PORT,
        'ROUTE_NAME': ROUTE_NAME,
        'VEHICLE_NAME': VEHICLE_NAME,
        'DOOR_NUMBER': DOOR_NUMBER,
        'STOP_LIST': STOP_LIST,
        'COUNT_POINT_MODE': COUNT_POINT_MODE,
        'REVERSE_COUNT_DIRECTION': REVERSE_COUNT_DIRECTION,
        'CAMERA_ROTATION': CAMERA_ROTATION,
        'YOLO_CONFIDENCE': YOLO_CONFIDENCE,
        'PASSENGER_ID_THRESHOLD': PASSENGER_ID_THRESHOLD,
        'YOLO_EVERY_N_FRAMES': YOLO_EVERY_N_FRAMES,
        'YOLO_IMGSZ': YOLO_IMGSZ,
        'MIN_AREA': MIN_AREA,
        'TRACKING_MAX_DIST_RATIO': TRACKING_MAX_DIST_RATIO,
        'TRACKING_LOST_FRAMES': TRACKING_LOST_FRAMES,
        'LINE_ORIENTATION': LINE_ORIENTATION,
        'LINE_Y': LINE_Y, 'LINE_X': LINE_X,
        'LINE_START_X': LINE_START_X, 'LINE_END_X': LINE_END_X,
        'LINE_START_Y': LINE_START_Y, 'LINE_END_Y': LINE_END_Y,
        'current_stop_index': current_stop_index,
        'TIME_AT_STOP_SEC': TIME_AT_STOP_SEC,  # время стоянки для выявления транзитов
    }
    try:
        with open(CONFIG_PATH, 'w', encoding='utf-8') as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

def load_config():
    """День 11: загружает пороги из JSON, если файл существует."""
    global _config_loaded
    global VIDEO_SOURCE, PHONE_IP, PHONE_PORT
    global ROUTE_NAME, VEHICLE_NAME, DOOR_NUMBER
    global COUNT_POINT_MODE, REVERSE_COUNT_DIRECTION, CAMERA_ROTATION
    global YOLO_CONFIDENCE, PASSENGER_ID_THRESHOLD, YOLO_EVERY_N_FRAMES, YOLO_IMGSZ
    global MIN_AREA, TRACKING_MAX_DIST_RATIO, TRACKING_LOST_FRAMES
    global LINE_ORIENTATION, LINE_Y, LINE_X
    global LINE_START_X, LINE_END_X, LINE_START_Y, LINE_END_Y
    global current_stop_index, STOP_NAME, TIME_AT_STOP_SEC
    if not os.path.isfile(CONFIG_PATH):
        return
    try:
        with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
            cfg = json.load(f)
        VIDEO_SOURCE = cfg.get('VIDEO_SOURCE', VIDEO_SOURCE)
        PHONE_IP = cfg.get('PHONE_IP', PHONE_IP)
        PHONE_PORT = int(cfg.get('PHONE_PORT', PHONE_PORT) or PHONE_PORT)
        ROUTE_NAME = cfg.get('ROUTE_NAME', ROUTE_NAME)
        VEHICLE_NAME = cfg.get('VEHICLE_NAME', VEHICLE_NAME)
        DOOR_NUMBER = int(cfg.get('DOOR_NUMBER', DOOR_NUMBER) or DOOR_NUMBER)
        COUNT_POINT_MODE = cfg.get('COUNT_POINT_MODE', COUNT_POINT_MODE)
        REVERSE_COUNT_DIRECTION = bool(cfg.get('REVERSE_COUNT_DIRECTION', REVERSE_COUNT_DIRECTION))
        CAMERA_ROTATION = int(cfg.get('CAMERA_ROTATION', CAMERA_ROTATION) or CAMERA_ROTATION)
        loaded_stops = cfg.get('STOP_LIST', STOP_LIST)
        if isinstance(loaded_stops, list) and loaded_stops:
            apply_transport_settings(stop_list=';'.join(str(s) for s in loaded_stops))
        YOLO_CONFIDENCE = cfg.get('YOLO_CONFIDENCE', YOLO_CONFIDENCE)
        PASSENGER_ID_THRESHOLD = cfg.get('PASSENGER_ID_THRESHOLD', PASSENGER_ID_THRESHOLD)
        YOLO_EVERY_N_FRAMES = cfg.get('YOLO_EVERY_N_FRAMES', YOLO_EVERY_N_FRAMES)
        YOLO_IMGSZ = cfg.get('YOLO_IMGSZ', YOLO_IMGSZ)
        MIN_AREA = int(cfg.get('MIN_AREA', MIN_AREA) or MIN_AREA)
        TRACKING_MAX_DIST_RATIO = float(cfg.get('TRACKING_MAX_DIST_RATIO', TRACKING_MAX_DIST_RATIO) or TRACKING_MAX_DIST_RATIO)
        TRACKING_LOST_FRAMES = int(cfg.get('TRACKING_LOST_FRAMES', TRACKING_LOST_FRAMES) or TRACKING_LOST_FRAMES)
        LINE_ORIENTATION = cfg.get('LINE_ORIENTATION', LINE_ORIENTATION)
        LINE_Y = cfg.get('LINE_Y', LINE_Y)
        LINE_X = cfg.get('LINE_X', LINE_X)
        LINE_START_X = cfg.get('LINE_START_X', LINE_START_X)
        LINE_END_X = cfg.get('LINE_END_X', LINE_END_X)
        LINE_START_Y = cfg.get('LINE_START_Y', LINE_START_Y)
        LINE_END_Y = cfg.get('LINE_END_Y', LINE_END_Y)
        idx = cfg.get('current_stop_index', current_stop_index)
        if 0 <= idx < len(STOP_LIST):
            current_stop_index = idx
            STOP_NAME = STOP_LIST[idx]
            load_stop_counters(STOP_NAME)
        TIME_AT_STOP_SEC = cfg.get('TIME_AT_STOP_SEC', TIME_AT_STOP_SEC)
        _config_loaded = True
        print(f'Конфигурация загружена из {CONFIG_PATH}')
    except Exception as e:
        print(f'Ошибка загрузки конфигурации: {e}')


def sync_source_settings_from_source(source):
    global VIDEO_SOURCE, PHONE_IP, PHONE_PORT, PHONE_STREAM
    VIDEO_SOURCE = source
    if isinstance(source, str):
        src = source.strip()
        if src.startswith(('http://', 'https://')):
            parsed = urlparse(src)
            if parsed.hostname:
                PHONE_IP = parsed.hostname
            if parsed.port:
                PHONE_PORT = int(parsed.port)
            PHONE_STREAM = src


def apply_counting_settings(point_mode=None, reverse_direction=None):
    global COUNT_POINT_MODE, REVERSE_COUNT_DIRECTION

    if point_mode in ('center', 'bottom'):
        COUNT_POINT_MODE = point_mode
    if reverse_direction is not None:
        REVERSE_COUNT_DIRECTION = bool(reverse_direction)


def get_count_anchor_point(x, y, w, h):
    if COUNT_POINT_MODE == 'bottom':
        return int(x + w / 2), int(y + h * 0.85)
    return int(x + w / 2), int(y + h / 2)


def resolve_count_event(raw_event):
    if REVERSE_COUNT_DIRECTION:
        return 'out' if raw_event == 'in' else 'in'
    return raw_event


def apply_camera_rotation(frame):
    if CAMERA_ROTATION == 90:
        return cv2.rotate(frame, cv2.ROTATE_90_CLOCKWISE)
    if CAMERA_ROTATION == 180:
        return cv2.rotate(frame, cv2.ROTATE_180)
    if CAMERA_ROTATION == 270:
        return cv2.rotate(frame, cv2.ROTATE_90_COUNTERCLOCKWISE)
    return frame

def _signal_handler(signum, _frame):
    """День 11: обработчик SIGINT/SIGTERM для корректного завершения."""
    global _shutdown_requested
    _shutdown_requested = True
    print('\n[Watchdog] Получен сигнал завершения, сохраняем данные...')

signal.signal(signal.SIGINT, _signal_handler)
signal.signal(signal.SIGTERM, _signal_handler)

class WatchdogThread(threading.Thread):
    """День 11: фоновый поток-сторож, отслеживает зависание основного цикла.

    Если основной цикл не обновляет heartbeat дольше timeout секунд,
    сторож сохраняет конфигурацию и счётчики (graceful degradation).
    """
    def __init__(self, timeout_sec=15.0):
        super().__init__(daemon=True, name='Watchdog')
        self._timeout = timeout_sec
        self._heartbeat = time.time()
        self._running = True

    def ping(self):
        """Вызывается из основного цикла каждый кадр."""
        self._heartbeat = time.time()

    def stop(self):
        self._running = False

    def run(self):
        while self._running:
            time.sleep(2.0)
            elapsed = time.time() - self._heartbeat
            if elapsed > self._timeout:
                print(f'[Watchdog] Основной цикл не отвечает {elapsed:.0f}s — сохраняем состояние')
                try:
                    save_current_stop_counters()
                    save_config()
                except Exception:
                    pass


# ── День 12: Профилирование и оптимизация памяти ────────────────────────────
# Модуль для замера потребления ОЗУ, количества кэшированных дескрипторов,
# размера пула треков и общего состояния системы. Данные выводятся по
# клавише 'P' и добавляются в веб-дашборд.

class PerformanceProfiler:
    """День 12: сборщик статистики производительности и потребления памяти.

    Отслеживает:
    - Потребление ОЗУ текущим процессом (через os / psutil если есть)
    - Количество активных треков и кэшированных дескрипторов
    - Среднее время цикла за последние N кадров
    - Статистику по FPS (мин/макс/среднее)
    """

    def __init__(self, window_size=120):
        # Размер скользящего окна для статистики (120 кадров ≈ 4 сек при 30fps)
        self._window = window_size
        # История времён цикла (секунды)
        self._loop_times = []
        # Счётчики пиковых значений
        self._peak_mem_mb = 0.0
        self._peak_tracks = 0
        self._total_frames = 0
        self._start_time = time.time()

    def record_loop(self, loop_sec: float, num_tracks: int):
        """Записывает время одного цикла и число активных треков."""
        self._loop_times.append(loop_sec)
        # Обрезаем историю до размера окна
        if len(self._loop_times) > self._window:
            self._loop_times = self._loop_times[-self._window:]
        self._total_frames += 1
        # Обновляем пиковое число треков
        if num_tracks > self._peak_tracks:
            self._peak_tracks = num_tracks

    def get_memory_mb(self) -> float:
        """Возвращает потребление ОЗУ текущим процессом в МБ."""
        try:
            # Попытка через psutil (если установлен)
            import psutil
            proc = psutil.Process(os.getpid())
            mem = proc.memory_info().rss / (1024 * 1024)
        except ImportError:
            # Fallback: чтение из /proc или Windows API
            try:
                import ctypes
                kernel32 = ctypes.windll.kernel32
                class PROCESS_MEMORY_COUNTERS(ctypes.Structure):
                    _fields_ = [
                        ("cb", ctypes.c_ulong),
                        ("PageFaultCount", ctypes.c_ulong),
                        ("PeakWorkingSetSize", ctypes.c_size_t),
                        ("WorkingSetSize", ctypes.c_size_t),
                        ("QuotaPeakPagedPoolUsage", ctypes.c_size_t),
                        ("QuotaPagedPoolUsage", ctypes.c_size_t),
                        ("QuotaPeakNonPagedPoolUsage", ctypes.c_size_t),
                        ("QuotaNonPagedPoolUsage", ctypes.c_size_t),
                        ("PagefileUsage", ctypes.c_size_t),
                        ("PeakPagefileUsage", ctypes.c_size_t),
                    ]
                pmc = PROCESS_MEMORY_COUNTERS()
                pmc.cb = ctypes.sizeof(pmc)
                handle = ctypes.windll.kernel32.GetCurrentProcess()
                ctypes.windll.psapi.GetProcessMemoryInfo(
                    handle, ctypes.byref(pmc), pmc.cb
                )
                mem = pmc.WorkingSetSize / (1024 * 1024)
            except Exception:
                mem = 0.0
        # Отслеживаем пик
        if mem > self._peak_mem_mb:
            self._peak_mem_mb = mem
        return mem

    def get_stats(self, num_tracks: int = 0, num_cached: int = 0) -> dict:
        """Возвращает словарь со всей статистикой."""
        mem = self.get_memory_mb()
        avg_loop = sum(self._loop_times) / max(1, len(self._loop_times))
        fps_avg = 1.0 / max(avg_loop, 0.001)
        fps_min = 1.0 / max(max(self._loop_times) if self._loop_times else 1.0, 0.001)
        fps_max = 1.0 / max(min(self._loop_times) if self._loop_times else 1.0, 0.001)
        uptime = time.time() - self._start_time
        return {
            'mem_mb': round(mem, 1),
            'peak_mem_mb': round(self._peak_mem_mb, 1),
            'avg_loop_ms': round(avg_loop * 1000, 1),
            'fps_avg': round(fps_avg, 1),
            'fps_min': round(fps_min, 1),
            'fps_max': round(fps_max, 1),
            'active_tracks': num_tracks,
            'peak_tracks': self._peak_tracks,
            'cached_descriptors': num_cached,
            'total_frames': self._total_frames,
            'uptime_sec': round(uptime, 0),
        }

    def format_report(self, num_tracks: int = 0, num_cached: int = 0) -> str:
        """День 12: форматирует отчёт для вывода в консоль/на панель."""
        s = self.get_stats(num_tracks, num_cached)
        uptime_min = s['uptime_sec'] / 60.0
        return (
            f"\n{'='*55}\n"
            f"  ПРОФИЛЬ ПРОИЗВОДИТЕЛЬНОСТИ (День 12)\n"
            f"{'='*55}\n"
            f"  ОЗУ:    {s['mem_mb']} МБ  (пик: {s['peak_mem_mb']} МБ)\n"
            f"  FPS:    avg={s['fps_avg']}  min={s['fps_min']}  max={s['fps_max']}\n"
            f"  Цикл:   {s['avg_loop_ms']} мс (среднее)\n"
            f"  Треки:  {s['active_tracks']} активных  (пик: {s['peak_tracks']})\n"
            f"  Кэш:    {s['cached_descriptors']} дескрипторов\n"
            f"  Кадров: {s['total_frames']}  |  Uptime: {uptime_min:.1f} мин\n"
            f"{'='*55}"
        )


# ── День 12: Очистка устаревших треков (оптимизация памяти) ──────────────────
# Треки, потерянные более чем TRACK_CLEANUP_FRAMES кадров назад, удаляются
# из словарей tracked_objects, track_pids, track_id_frame, appearance_frame.
# Это предотвращает утечку памяти при длительной работе.
TRACK_CLEANUP_FRAMES = 300   # каждые 300 кадров чистим мёртвые треки
TRACK_MAX_LOST_AGE = 60      # удаляем треки, потерянные > 60 кадров назад

def cleanup_stale_tracks(
    tracked_objects: dict, track_pids: dict,
    track_id_frame: dict, appearance_frame: dict,
    frame_idx: int
):
    """День 12: удаляет давно потерянные треки для экономии ОЗУ.

    Вызывается каждые TRACK_CLEANUP_FRAMES кадров из основного цикла.
    Удаляет объекты, у которых lost_count (td[3]) >= TRACK_MAX_LOST_AGE.
    Формат tracked_objects: {id: (center, crossed, direction, lost_frames)}
    """
    stale_ids = []
    for obj_id, td in tracked_objects.items():
        lost = td[3] if len(td) > 3 else 0
        if lost >= TRACK_MAX_LOST_AGE:
            stale_ids.append(obj_id)
    # Удаляем из всех словарей
    for sid in stale_ids:
        tracked_objects.pop(sid, None)
        track_pids.pop(sid, None)
        track_id_frame.pop(sid, None)
        appearance_frame.pop(sid, None)
    if stale_ids:
        print(f'[Cleanup] Удалено {len(stale_ids)} устаревших треков (frame {frame_idx})')
    return len(stale_ids)


# ── День 13: Скрипт сборки EXE через PyInstaller ────────────────────────────
# Генерирует обновлённый .spec файл и batch-скрипт для сборки.
# Включает: CPU-only torch, все модули, yolov8n.pt, шрифты, данные ultralytics.

def generate_build_script():
    """День 13: создаёт build_exe.bat и обновлённый People_Counter_v2.spec.

    Скрипт:
    1. Устанавливает CPU-only torch (без CUDA — экономия ~1.5 ГБ)
    2. Собирает EXE через PyInstaller с нужными hidden imports
    3. Копирует yolov8n.pt и конфигурацию в dist/
    """
    spec_content = '''# -*- mode: python ; coding: utf-8 -*-
# ── День 13: обновлённый spec-файл для PyInstaller ──────────────────────
# Включает все модули проекта, данные ultralytics, модель YOLO,
# шрифты для корректного отображения кириллицы.

from PyInstaller.utils.hooks import collect_data_files, collect_submodules

# Собираем все данные ultralytics (конфиги YOLO, yaml-файлы)
datas = collect_data_files('ultralytics')
# Добавляем модель YOLO и вспомогательные модули
datas += [
    ('yolov8n.pt', '.'),                        # модель нейросети
    ('passenger_identifier.py', '.'),            # идентификатор пассажиров
    ('od_matrix_tracker.py', '.'),               # OD-матрица
    ('cross_route_stitcher.py', '.'),            # сшивка маршрутов
]

# Скрытые импорты: модули, которые PyInstaller не находит автоматически
hidden = collect_submodules('ultralytics')
hidden += [
    'mediapipe',
    'mediapipe.python',
    'mediapipe.python._framework_bindings',
    'PIL', 'PIL.Image', 'PIL.ImageDraw', 'PIL.ImageFont',
    'pandas', 'openpyxl', 'numpy',
    'sqlite3',
    'passenger_identifier', 'od_matrix_tracker', 'cross_route_stitcher',
]

a = Analysis(
    ['people_counter.py'],
    pathex=['.'],
    binaries=[],
    datas=datas,
    hiddenimports=hidden,
    hookspath=[],
    hooksconfig={},
    runtime_hooks=[],
    # Исключаем ненужные модули для уменьшения размера
    excludes=[
        'tkinter', 'matplotlib', 'scipy', 'IPython',
        'jupyter', 'notebook', 'pytest',
        'torch.cuda', 'torchvision',  # CPU-only: исключаем CUDA
    ],
    noarchive=False,
    optimize=1,  # байткод-оптимизация
)

pyz = PYZ(a.pure)

exe = EXE(
    pyz,
    a.scripts,
    [],
    exclude_binaries=True,
    name='People_Counter_v2',
    debug=False,
    bootloader_ignore_signals=False,
    strip=False,
    upx=True,                    # сжатие UPX (уменьшает размер)
    console=True,                # консоль для вывода диагностики
    disable_windowed_traceback=False,
    argv_emulation=False,
    icon=None,                   # можно указать .ico файл
)

coll = COLLECT(
    exe,
    a.binaries,
    a.datas,
    strip=False,
    upx=True,
    upx_exclude=[],
    name='People_Counter_v2',
)
'''

    # Batch-скрипт для сборки
    bat_content = '''@echo off
chcp 65001 >nul
echo ═══════════════════════════════════════════════════════
echo   День 13: Сборка People_Counter_v2.exe
echo   НМУ ВКС — Определение траекторий движения пассажиров
echo ═══════════════════════════════════════════════════════
echo.

REM Шаг 1: Установка CPU-only PyTorch (экономия ~1.5 ГБ без CUDA)
echo [1/4] Установка CPU-only PyTorch...
pip install torch torchvision --index-url https://download.pytorch.org/whl/cpu --quiet 2>nul
echo       Готово.

REM Шаг 2: Установка зависимостей
echo [2/4] Проверка зависимостей...
pip install pyinstaller ultralytics mediapipe opencv-python pandas openpyxl pillow --quiet 2>nul
echo       Готово.

REM Шаг 3: Сборка через PyInstaller
echo [3/4] Сборка EXE через PyInstaller...
pyinstaller People_Counter_v2.spec --noconfirm --clean
if %errorlevel% neq 0 (
    echo ОШИБКА: Сборка не удалась!
    pause
    exit /b 1
)
echo       Готово.

REM Шаг 4: Копируем дополнительные файлы в dist/
echo [4/4] Копирование файлов...
if exist "yolov8n.pt" copy /Y "yolov8n.pt" "dist\\People_Counter_v2\\" >nul
if exist "people_counter_config.json" copy /Y "people_counter_config.json" "dist\\People_Counter_v2\\" >nul
echo       Готово.

echo.
echo ═══════════════════════════════════════════════════════
echo   Сборка завершена!
echo   EXE: dist\\People_Counter_v2\\People_Counter_v2.exe
echo ═══════════════════════════════════════════════════════
pause
'''

    try:
        with open('People_Counter_v2.spec', 'w', encoding='utf-8') as f:
            f.write(spec_content)
        with open('build_exe_v2.bat', 'w', encoding='utf-8') as f:
            f.write(bat_content)
        print('День 13: Файлы сборки созданы:')
        print('  • People_Counter_v2.spec')
        print('  • build_exe_v2.bat')
        print('Для сборки запустите: build_exe_v2.bat')
    except Exception as e:
        print(f'Ошибка создания файлов сборки: {e}')


# ── День 14: Генерация пользовательской документации ─────────────────────────
# Автоматическое создание файла ИНСТРУКЦИЯ_v2.txt с полным описанием
# системы, установки, использования, горячих клавиш и устранения неполадок.

def generate_user_manual():
    """День 14: генерирует полную инструкцию пользователя.

    Создаёт файл ИНСТРУКЦИЯ_ПРОГ2.txt с разделами:
      1. Назначение системы
      2. Системные требования
      3. Установка и первый запуск
      4. Интерфейс и управление
      5. Настройки и конфигурация
      6. Отчёты и экспорт
      7. Веб-дашборд
      8. Сборка EXE
      9. Устранение неполадок
    """
    manual = '''
╔════════════════════════════════════════════════════════════════════════════╗
║        ИНСТРУКЦИЯ ПОЛЬЗОВАТЕЛЯ — ПРОГРАММА 2 (v2.0)                      ║
║        «Определение траекторий движения пассажиров»                       ║
║        НМУ ВКС, Воронеж                                                  ║
╚════════════════════════════════════════════════════════════════════════════╝

════════════════════════════════════════════════════════════════════════════
 1. НАЗНАЧЕНИЕ СИСТЕМЫ
════════════════════════════════════════════════════════════════════════════

Система предназначена для автоматического подсчёта и идентификации
пассажиров в общественном транспорте с помощью видеокамеры.

Основные возможности:
 • Детекция людей нейросетью YOLOv8n в реальном времени
 • Идентификация пассажиров по параметрам лица (214 точек)
 • Распознавание по внешности (рост, цвета одежды)
 • Матрица корреспонденций «откуда-куда» (OD-матрица)
 • Построение траекторий движения по маршруту
 • Сшивка поездок через несколько маршрутов (пересадки)
 • Выявление транзитных пассажиров (вход/выход на одной остановке)
 • Веб-дашборд для удалённого мониторинга (порт 8888)
 • Экспорт отчётов: Excel (12 листов), CSV, JSON

════════════════════════════════════════════════════════════════════════════
 2. СИСТЕМНЫЕ ТРЕБОВАНИЯ
════════════════════════════════════════════════════════════════════════════

Минимальные:
 • ОС: Windows 10/11 (64-bit)
 • Процессор: Intel Core i5 / AMD Ryzen 5 (4 ядра)
 • ОЗУ: 8 ГБ
 • Диск: 2 ГБ свободного места
 • Камера: USB-камера, IP-камера (RTSP/HTTP) или видеофайл

Рекомендуемые:
 • Процессор: Intel Core i7 / AMD Ryzen 7 (8 ядер)
 • ОЗУ: 16 ГБ
 • SSD-диск
 • GPU: NVIDIA с поддержкой CUDA (ускорение YOLO в 3-5 раз)

Программные зависимости (при запуске из исходников):
 • Python 3.10+
 • opencv-python, ultralytics, mediapipe
 • numpy, pandas, openpyxl, pillow
 • torch (CPU-only достаточно)

════════════════════════════════════════════════════════════════════════════
 3. УСТАНОВКА И ПЕРВЫЙ ЗАПУСК
════════════════════════════════════════════════════════════════════════════

Вариант А — из EXE (рекомендуется):
 1. Распакуйте архив People_Counter_v2.zip
 2. Запустите People_Counter_v2.exe
 3. Выберите источник видео в диалоге

Вариант Б — из исходников:
 1. Установите Python 3.10+
 2. Установите зависимости:
      pip install opencv-python ultralytics mediapipe
      pip install numpy pandas openpyxl pillow
      pip install torch --index-url https://download.pytorch.org/whl/cpu
 3. Запустите:
      python people_counter.py

При первом запуске модель yolov8n.pt скачается автоматически (~6 МБ).

════════════════════════════════════════════════════════════════════════════
 4. ИНТЕРФЕЙС И УПРАВЛЕНИЕ
════════════════════════════════════════════════════════════════════════════

Основное окно:
 • Верхняя часть — видеопоток с наложенными боксами и ID пассажиров
 • Линия подсчёта — зелёная/красная линия, пересечение = вход/выход
 • Нижняя панель — счётчики, FPS, название остановки

Горячие клавиши:
 ┌─────────┬──────────────────────────────────────────────┐
 │ Клавиша │ Действие                                     │
 ├─────────┼──────────────────────────────────────────────┤
 │ Q / Й   │ Выход из программы                           │
 │ H / Р   │ Горизонтальная линия подсчёта                │
 │ V / М   │ Вертикальная линия подсчёта                  │
 │ C / С   │ Сброс счётчиков, перерисовка линии           │
 │ PgUp /< │ Предыдущая остановка                         │
 │ PgDn /> │ Следующая остановка                          │
 │ S / Ы   │ Окно управления остановками                  │
 │ R / К   │ Генерация всех отчётов                       │
 │ X / Ч   │ Запуск ночной сшивки маршрутов               │
 │ P / З   │ Профиль производительности (ОЗУ, FPS)        │
 │ B / И   │ Создать файлы сборки EXE                     │
 │ [ / ]   │ YOLO_CONFIDENCE ±0.05  (порог детекции)      │
 │ - / =   │ ID_THRESHOLD ±0.02  (порог идентификации)    │
 │ ; / '   │ TIME_AT_STOP_SEC ±5  (время стоянки)         │
 └─────────┴──────────────────────────────────────────────┘

Линию подсчёта также можно нарисовать мышкой (зажать ЛКМ и провести).

════════════════════════════════════════════════════════════════════════════
 5. НАСТРОЙКИ И КОНФИГУРАЦИЯ
════════════════════════════════════════════════════════════════════════════

Файл конфигурации: people_counter_config.json
Создаётся автоматически при выходе из программы. Содержит:
 • YOLO_CONFIDENCE  — порог уверенности детекции (0.10–0.95)
 • PASSENGER_ID_THRESHOLD — порог идентификации (0.05–0.60)
 • TIME_AT_STOP_SEC — время стоянки для транзитов (5–300 сек)
 • Положение линии подсчёта
 • Индекс текущей остановки

Параметры в начале файла people_counter.py:
 • VIDEO_SOURCE — источник видео (камера / файл / URL)
 • STOP_LIST — список остановок маршрута
 • ROUTE_NAME, VEHICLE_NAME — название маршрута и ТС
 • YOLO_MODEL — модель YOLO (yolov8n/s/m)
 • DISPLAY_W, DISPLAY_H — размер окна видео

════════════════════════════════════════════════════════════════════════════
 6. ОТЧЁТЫ И ЭКСПОРТ
════════════════════════════════════════════════════════════════════════════

Нажмите R для генерации. Файлы сохраняются в папку reports/:

CSV-отчёты:
 • daily_report.csv — по дням
 • hourly_report.csv — по часам
 • stop_report.csv — по остановкам
 • trajectory_report.csv — траектории пассажиров
 • appearance_report.csv — параметры внешности
 • transit_passengers_report.csv — транзитные пассажиры
 • correspondence_matrix.csv — матрица корреспонденций

Excel-отчёт (passenger_flow_report.xlsx):
 12 листов с цветным оформлением:
 • Все данные, По дням, По часам, По месяцам, По годам
 • По остановкам, По маршрутам, Панель руководителя
 • Матрица корреспонд., Траектории, Внешность
 • Транзитные пассажиры, Сшивка маршрутов, OD сводка, Время в пути

═══════════════════════════════════════════════════════════════════════════
 7. ВЕБ-ДАШБОРД
════════════════════════════════════════════════════════════════════════════

Автоматически запускается на порту 8888.
Откройте в браузере: http://localhost:8888

Отображает в реальном времени:
 • Счётчики входа/выхода/в салоне
 • Текущий маршрут и остановку
 • OD-матрицу корреспонденций
 • Таблицу пассажиров с ID и параметрами

════════════════════════════════════════════════════════════════════════════
 8. СБОРКА EXE (для распространения без Python)
════════════════════════════════════════════════════════════════════════════

1. Нажмите B в работающей программе — создадутся файлы сборки
2. Запустите build_exe_v2.bat
3. Готовый EXE будет в папке dist/People_Counter_v2/

Размер EXE: ~250–350 МБ (включает torch CPU-only, ultralytics, mediapipe)

════════════════════════════════════════════════════════════════════════════
 9. УСТРАНЕНИЕ НЕПОЛАДОК
════════════════════════════════════════════════════════════════════════════

Проблема: «ultralytics не установлен»
 → pip install ultralytics

Проблема: «mediapipe не установлен» (идентификация отключена)
 → pip install mediapipe

Проблема: Камера не открывается
 → Проверьте VIDEO_SOURCE (0=USB, URL=сетевая)
 → Разрешите доступ к камере в настройках Windows

Проблема: Низкий FPS (<10)
 → Увеличьте YOLO_EVERY_N_FRAMES до 3–4
 → Уменьшите YOLO_IMGSZ до 320
 → Используйте YOLO_HALF=True на GPU

Проблема: Много ложных идентификаций
 → Уменьшите PASSENGER_ID_THRESHOLD (клавиша -)
 → Рекомендуемый диапазон: 0.15–0.25

Проблема: Пропуск пассажиров
 → Уменьшите YOLO_CONFIDENCE (клавиша [)
 → Проверьте расположение линии подсчёта

════════════════════════════════════════════════════════════════════════════
 КОНТАКТЫ
════════════════════════════════════════════════════════════════════════════

 Организация: НМУ ВКС
 Телефон:     +7 952 553-96-21
 Email:       mveo@yandex.ru
 Город:       Воронеж

════════════════════════════════════════════════════════════════════════════
'''

    path = 'ИНСТРУКЦИЯ_ПРОГ2.txt'
    try:
        with open(path, 'w', encoding='utf-8') as f:
            f.write(manual.strip())
        print(f'День 14: Инструкция пользователя создана: {path}')
    except Exception as e:
        print(f'Ошибка создания инструкции: {e}')


# ── День 15: Финальная самодиагностика системы ───────────────────────────────
# Проверяет все компоненты перед штатной работой: модули, БД, модель,
# шрифты, директории. Выводит таблицу готовности.

def run_self_diagnostic() -> bool:
    """День 15: полная самодиагностика перед запуском.

    Проверяет:
      ✔ Наличие всех Python-модулей (opencv, ultralytics, mediapipe, pandas...)
      ✔ Доступность модели YOLO (yolov8n.pt)
      ✔ Возможность создания/записи БД
      ✔ Создание папки отчётов
      ✔ Доступность шрифта для кириллицы
      ✔ Наличие вспомогательных модулей (passenger_identifier, od_matrix_tracker)

    Возвращает True если все критические компоненты в порядке.
    """
    print()
    print('╔══════════════════════════════════════════════════╗')
    print('║   САМОДИАГНОСТИКА СИСТЕМЫ (День 15)             ║')
    print('╚══════════════════════════════════════════════════╝')

    # Список проверок: (название, условие, критичность)
    checks = []
    all_ok = True

    # 1. OpenCV
    try:
        import cv2 as _cv2
        ver = _cv2.__version__
        checks.append(('OpenCV', f'✔ v{ver}', True))
    except ImportError:
        checks.append(('OpenCV', '✘ НЕ УСТАНОВЛЕН', False))
        all_ok = False

    # 2. NumPy
    try:
        import numpy as _np
        checks.append(('NumPy', f'✔ v{_np.__version__}', True))
    except ImportError:
        checks.append(('NumPy', '✘ НЕ УСТАНОВЛЕН', False))
        all_ok = False

    # 3. Ultralytics (YOLO)
    try:
        from ultralytics import YOLO as _Y
        import ultralytics
        checks.append(('Ultralytics', f'✔ v{ultralytics.__version__}', True))
    except ImportError:
        checks.append(('Ultralytics', '✘ НЕ УСТАНОВЛЕН', False))
        all_ok = False

    # 4. PyTorch
    try:
        import torch
        gpu = 'CUDA' if torch.cuda.is_available() else 'CPU-only'
        checks.append(('PyTorch', f'✔ v{torch.__version__} ({gpu})', True))
    except ImportError:
        checks.append(('PyTorch', '⚠ не установлен (YOLO может работать без)', True))

    # 5. MediaPipe
    try:
        import mediapipe as _mp
        checks.append(('MediaPipe', f'✔ v{_mp.__version__}', True))
    except ImportError:
        checks.append(('MediaPipe', '⚠ не установлен (идентификация отключена)', True))

    # 6. Pandas
    if pd is not None:
        checks.append(('Pandas', f'✔ v{pd.__version__}', True))
    else:
        checks.append(('Pandas', '⚠ не установлен (отчёты отключены)', True))

    # 7. Openpyxl
    try:
        import openpyxl
        checks.append(('Openpyxl', f'✔ v{openpyxl.__version__}', True))
    except ImportError:
        checks.append(('Openpyxl', '⚠ не установлен (Excel-отчёты отключены)', True))

    # 8. Pillow
    if Image is not None:
        import PIL
        checks.append(('Pillow', f'✔ v{PIL.__version__}', True))
    else:
        checks.append(('Pillow', '⚠ не установлен (Unicode-текст ограничен)', True))

    # 9. Модель YOLO
    if os.path.isfile(YOLO_MODEL):
        size_mb = os.path.getsize(YOLO_MODEL) / (1024 * 1024)
        checks.append(('Модель YOLO', f'✔ {YOLO_MODEL} ({size_mb:.1f} МБ)', True))
    else:
        checks.append(('Модель YOLO', f'⚠ {YOLO_MODEL} не найден (скачается автоматически)', True))

    # 10. Модуль идентификации
    checks.append(('PassengerIdentifier',
                    '✔ загружен' if IDENT_AVAILABLE else '⚠ не загружен',
                    True))

    # 11. Модуль OD-матрицы
    checks.append(('ODMatrixTracker',
                    '✔ загружен' if OD_AVAILABLE else '⚠ не загружен',
                    True))

    # 12. Директория отчётов
    try:
        os.makedirs(REPORT_FOLDER, exist_ok=True)
        checks.append(('Папка отчётов', f'✔ {REPORT_FOLDER}/', True))
    except Exception:
        checks.append(('Папка отчётов', f'✘ не удалось создать {REPORT_FOLDER}/', False))

    # 13. БД
    try:
        _c = sqlite3.connect(DATABASE_PATH)
        _c.execute('SELECT 1')
        _c.close()
        checks.append(('База данных', f'✔ {DATABASE_PATH}', True))
    except Exception:
        checks.append(('База данных', f'✘ ошибка доступа к {DATABASE_PATH}', False))
        all_ok = False

    # 14. Конфигурация
    if os.path.isfile(CONFIG_PATH):
        checks.append(('Конфигурация', f'✔ {CONFIG_PATH} (загружена)', True))
    else:
        checks.append(('Конфигурация', '— файл не найден (будут значения по умолчанию)', True))

    # Вывод таблицы
    print()
    print(f'  {"Компонент":<25} {"Статус"}')
    print(f'  {"─"*25} {"─"*40}')
    for name, status, ok in checks:
        print(f'  {name:<25} {status}')
    print()

    if all_ok:
        print('  ▶ Все критические компоненты готовы к работе.')
    else:
        print('  ▶ ВНИМАНИЕ: есть критические ошибки! Устраните их перед запуском.')
    print()
    return all_ok


# ── День 15: Итоговая сводка при завершении ──────────────────────────────────
# Расширенный отчёт при выходе: статистика сессии, пики, рекомендации.

def print_session_summary(profiler, p_db, od_tracker):
    """День 15: печатает подробную сводку сессии при завершении.

    Включает: время работы, обработано кадров, пики ОЗУ/треков,
    количество идентифицированных пассажиров, OD-записей, рекомендации.
    """
    stats = profiler.get_stats() if profiler else {}
    uptime_min = stats.get('uptime_sec', 0) / 60.0

    print()
    print('╔══════════════════════════════════════════════════════════╗')
    print('║          ИТОГОВАЯ СВОДКА СЕССИИ (День 15)               ║')
    print('╠══════════════════════════════════════════════════════════╣')
    print(f'║  Время работы:     {uptime_min:>8.1f} мин                        ║')
    print(f'║  Обработано кадров: {stats.get("total_frames", 0):>7}                          ║')
    print(f'║  ОЗУ (текущая):    {stats.get("mem_mb", 0):>7.1f} МБ                        ║')
    print(f'║  ОЗУ (пиковая):    {stats.get("peak_mem_mb", 0):>7.1f} МБ                        ║')
    print(f'║  FPS (среднее):    {stats.get("fps_avg", 0):>8.1f}                          ║')
    print(f'║  Пик треков:       {stats.get("peak_tracks", 0):>8}                          ║')

    # Статистика идентификации
    if p_db is not None:
        try:
            unique = p_db.get_unique_count()
            print(f'║  Уникальных пассажиров: {unique:>4}                          ║')
        except Exception:
            pass

    # Статистика OD
    if od_tracker is not None:
        try:
            dur = od_tracker.get_duration_stats()
            print(f'║  OD-записей:      {dur.get("total_trips", 0):>8}                          ║')
        except Exception:
            pass

    print('╠══════════════════════════════════════════════════════════╣')

    # Рекомендации по результатам сессии
    recommendations = []
    if stats.get('fps_avg', 30) < 15:
        recommendations.append('  FPS ниже 15 — увеличьте YOLO_EVERY_N_FRAMES')
    if stats.get('peak_mem_mb', 0) > 2000:
        recommendations.append('  ОЗУ > 2 ГБ — уменьшите YOLO_IMGSZ или кэш треков')
    if stats.get('peak_tracks', 0) > 50:
        recommendations.append('  Много треков — проверьте MIN_AREA и YOLO_CONFIDENCE')
    if not recommendations:
        recommendations.append('  Параметры в норме. Рекомендаций нет.')
    print('║  Рекомендации:                                         ║')
    for r in recommendations:
        print(f'║  {r:<55} ║')
    print('╚══════════════════════════════════════════════════════════╝')


# ─── Веб-панель мониторинга ──────────────────────────────────────────────────
DASHBOARD_PORT = 8888
live_data = {
    'enter': 0, 'exit': 0, 'in_salon': 0,
    'stop': STOP_NAME, 'route': ROUTE_NAME,
    'vehicle': VEHICLE_NAME, 'door': DOOR_NUMBER,
    'time': '', 'fps': 0,
    'stop_index': 1, 'stop_total': len(STOP_LIST),
    'stops': STOP_LIST[:],
    'od_html': '',
    'pax_html': '',
}

DASHBOARD_HTML = '''<!DOCTYPE html>
<html lang="ru"><head><meta charset="utf-8">
<title>НМУ ВКС — Мониторинг</title>
<meta name="viewport" content="width=device-width,initial-scale=1">
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:'Segoe UI',Tahoma,sans-serif;background:#0d1117;color:#e6edf3;min-height:100vh;display:flex;flex-direction:column;align-items:center;padding:24px}
h1{color:#58a6ff;font-size:28px;margin-bottom:4px}
.sub{color:#8b949e;font-size:14px;margin-bottom:24px}
.cards{display:flex;gap:16px;flex-wrap:wrap;justify-content:center;margin-bottom:24px}
.card{background:#161b22;border:1px solid #30363d;border-radius:12px;padding:20px 32px;min-width:180px;text-align:center}
.card .label{font-size:13px;color:#8b949e;text-transform:uppercase;letter-spacing:1px;margin-bottom:8px}
.card .value{font-size:48px;font-weight:700}
.card.enter .value{color:#3fb950}
.card.exit .value{color:#d29922}
.card.salon .value{color:#f85149}
.card.total{border-color:#58a6ff}
.card.total .value{color:#58a6ff}
.info{background:#161b22;border:1px solid #30363d;border-radius:12px;padding:16px 28px;font-size:14px;color:#8b949e;line-height:1.8;min-width:300px;width:100%;max-width:700px}
.info span{color:#e6edf3}
.dot{display:inline-block;width:8px;height:8px;border-radius:50%;background:#3fb950;margin-right:6px;animation:pulse 1.5s infinite}
@keyframes pulse{0%,100%{opacity:1}50%{opacity:.3}}
.status{margin-top:16px;font-size:12px;color:#484f58}
table{width:100%;border-collapse:collapse;margin-top:8px}
th{background:#21262d;color:#8b949e;padding:8px 12px;text-align:left;font-size:12px;text-transform:uppercase;letter-spacing:0.5px}
td{padding:8px 12px;border-bottom:1px solid #21262d;font-size:14px}
tr.active{background:#1c2333}
tr:hover{background:#161b22}
.btn{background:#21262d;color:#e6edf3;border:1px solid #30363d;border-radius:6px;padding:4px 10px;cursor:pointer;font-size:12px}
.btn:hover{background:#30363d}
.btn-add{background:#238636;border-color:#2ea043;color:#fff;padding:6px 16px;font-size:13px;margin-top:8px}
.btn-add:hover{background:#2ea043}
.btn-del{color:#f85149;background:transparent;border:none;cursor:pointer;font-size:14px}
.btn-del:hover{text-decoration:underline}
input.ed{background:#0d1117;border:1px solid #30363d;color:#e6edf3;border-radius:4px;padding:4px 8px;font-size:13px;width:100%}
.nav-btn{background:#21262d;color:#e6edf3;border:1px solid #30363d;border-radius:8px;padding:6px 14px;font-size:18px;cursor:pointer}
</style></head><body>
<h1>НМУ ВКС</h1>
<p class="sub">Система подсчёта пассажиров — онлайн</p>
<div class="cards">
 <div class="card enter"><div class="label">Вошло (тек.)</div><div class="value" id="enter">0</div></div>
 <div class="card exit"><div class="label">Вышло (тек.)</div><div class="value" id="exit">0</div></div>
 <div class="card salon"><div class="label">В салоне</div><div class="value" id="salon">0</div></div>
</div>
<div class="cards">
 <div class="card total"><div class="label">Всего вошло</div><div class="value" id="totalIn">0</div></div>
 <div class="card total"><div class="label">Всего вышло</div><div class="value" id="totalOut">0</div></div>
 <div class="card total"><div class="label">В салоне итого</div><div class="value" id="totalSalon">0</div></div>
 <div class="card total" style="border-color:#60bfff66"><div class="label">Уник. пассажиры</div><div class="value" id="uniqPax" style="color:#60bfff">0</div></div>
</div>
<div class="info">
 <div><span class="dot"></span>Маршрут: <span id="route">—</span></div>
 <div>Транспорт: <span id="vehicle">—</span></div>
 <div>Остановка: <span id="stop">—</span> <span id="stopidx" style="color:#58a6ff"></span></div>
 <div>Дверь: <span id="door">—</span></div>
 <div>Время: <span id="time">—</span> | FPS: <span id="fps">—</span></div>
</div>
<div class="info" style="margin-top:12px">
 <div style="margin-bottom:8px;display:flex;align-items:center;gap:8px">
  <span style="color:#58a6ff;font-weight:600;flex:1">Остановки маршрута</span>
  <button class="nav-btn" onclick="chStop('prev')" id="btnPrev">◀</button>
  <button class="nav-btn" onclick="chStop('next')" id="btnNext">▶</button>
 </div>
 <table>
  <thead><tr><th>#</th><th>Остановка</th><th>Вошло</th><th>Вышло</th><th>Баланс</th><th></th></tr></thead>
  <tbody id="stopTable"></tbody>
  <tfoot><tr style="background:#21262d;font-weight:600"><td></td><td style="color:#58a6ff">Итого</td><td id="ftIn" style="color:#3fb950">0</td><td id="ftOut" style="color:#d29922">0</td><td id="ftBal" style="color:#f85149">0</td><td></td></tr></tfoot>
 </table>
 <div style="margin-top:10px;display:flex;gap:6px;align-items:center">
  <input class="ed" id="newStop" placeholder="Название новой остановки..." style="flex:1">
  <button class="btn btn-add" onclick="addStop()">➕ Добавить</button>
 </div>
</div>
<div class="info" style="margin-top:12px">
 <div style="margin-bottom:8px"><span style="color:#58a6ff;font-weight:600">OD-матрица (откуда-куда)</span></div>
 <div id="odMatrix" style="font-size:12px;overflow-x:auto">Загрузка...</div>
</div>
<div class="info" style="margin-top:12px">
 <div style="margin-bottom:8px"><span style="color:#58a6ff;font-weight:600">Таблица параметров пассажиров</span></div>
 <div id="paxTable" style="font-size:12px;overflow-x:auto">Загрузка...</div>
</div>
<div class="status" id="status">Подключение...</div>
<script>
function api(path,body){return fetch(path,{method:'POST',headers:{'Content-Type':'application/json'},body:body?JSON.stringify(body):undefined}).then(()=>update());}
function chStop(dir){api('/api/stop/'+dir);}
function addStop(){
 var n=document.getElementById('newStop').value.trim();
 if(n){api('/api/stop/add',{name:n});document.getElementById('newStop').value='';}
}
function renStop(i){
 var n=prompt('Новое название:');
 if(n&&n.trim())api('/api/stop/rename',{index:i,name:n.trim()});
}
function delStop(i){if(confirm('Удалить остановку?'))api('/api/stop/delete',{index:i});}
async function update(){
 try{
  const r=await fetch('/api/data');const d=await r.json();
  document.getElementById('enter').textContent=d.enter;
  document.getElementById('exit').textContent=d.exit;
  document.getElementById('salon').textContent=d.in_salon;
  document.getElementById('totalIn').textContent=d.total_enter;
  document.getElementById('totalOut').textContent=d.total_exit;
  document.getElementById('totalSalon').textContent=Math.max(0,d.total_enter-d.total_exit);
  document.getElementById('uniqPax').textContent=d.unique_passengers||0;
  if(d.od_html){var odarea=document.getElementById('odMatrix');if(odarea)odarea.innerHTML=d.od_html;}
  if(d.pax_html){var ptbl=document.getElementById('paxTable');if(ptbl)ptbl.innerHTML=d.pax_html;}
  document.getElementById('route').textContent=d.route;
  document.getElementById('vehicle').textContent=d.vehicle;
  document.getElementById('stop').textContent=d.stop;
  document.getElementById('stopidx').textContent='['+d.stop_index+'/'+d.stop_total+']';
  document.getElementById('door').textContent=d.door;
  document.getElementById('time').textContent=d.time;
  document.getElementById('fps').textContent=d.fps;
  var tb=document.getElementById('stopTable');tb.innerHTML='';
  var ss=d.stop_stats||[],sumIn=0,sumOut=0;
  ss.forEach(function(s,i){
   sumIn+=s.enter;sumOut+=s.exit;
   var tr=document.createElement('tr');
   if(i+1==d.stop_index)tr.className='active';
   tr.innerHTML='<td>'+(i+1)+'</td>'+
    '<td style="cursor:pointer" onclick="api(\'/api/stop/set\',{index:'+i+'})">'+(i+1==d.stop_index?'\u25b6 ':'')+s.name+'</td>'+
    '<td style="color:#3fb950">'+s.enter+'</td>'+
    '<td style="color:#d29922">'+s.exit+'</td>'+
    '<td style="color:'+(s.enter-s.exit>=0?'#58a6ff':'#f85149')+'">'+(s.enter-s.exit)+'</td>'+
    '<td><button class="btn" onclick="renStop('+i+')">\u270f</button> <button class="btn-del" onclick="delStop('+i+')">\u2716</button></td>';
   tb.appendChild(tr);
  });
  document.getElementById('ftIn').textContent=sumIn;
  document.getElementById('ftOut').textContent=sumOut;
  document.getElementById('ftBal').textContent=Math.max(0,sumIn-sumOut);
  document.getElementById('btnPrev').style.opacity=d.stop_index<=1?'0.3':'1';
  document.getElementById('btnNext').style.opacity=d.stop_index>=d.stop_total?'0.3':'1';
    document.getElementById('status').textContent='\u2705 '+(d.time||'');
 }catch(e){document.getElementById('status').textContent='\u274c Нет связи';}
}
setInterval(update,1000);update();
</script></body></html>'''


class DashboardHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        if self.path == '/api/od':
            self.send_response(200)
            self.send_header('Content-Type', 'text/html; charset=utf-8')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(live_data.get('od_html', '<p>Нет данных</p>').encode('utf-8'))
        elif self.path == '/api/passengers':
            self.send_response(200)
            self.send_header('Content-Type', 'text/html; charset=utf-8')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(live_data.get('pax_html', '<p>Нет данных</p>').encode('utf-8'))
        elif self.path == '/api/data':
            self.send_response(200)
            self.send_header('Content-Type', 'application/json; charset=utf-8')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(json.dumps(live_data, ensure_ascii=False).encode('utf-8'))
        else:
            self.send_response(200)
            self.send_header('Content-Type', 'text/html; charset=utf-8')
            self.send_header('Cache-Control', 'no-store, no-cache, must-revalidate, max-age=0')
            self.send_header('Pragma', 'no-cache')
            self.end_headers()
            self.wfile.write(DASHBOARD_HTML.encode('utf-8'))

    def do_POST(self):
        global current_stop_index, STOP_NAME
        length = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(length) if length else b''
        data = {}
        if body:
            try:
                data = json.loads(body)
            except Exception:
                pass

        if self.path == '/api/stop/prev':
            if current_stop_index > 0:
                switch_to_stop(current_stop_index - 1)
            self._json_ok()
        elif self.path == '/api/stop/next':
            if current_stop_index < len(STOP_LIST) - 1:
                switch_to_stop(current_stop_index + 1)
            self._json_ok()
        elif self.path == '/api/stop/set':
            idx = int(data.get('index', -1))
            if 0 <= idx < len(STOP_LIST):
                switch_to_stop(idx)
            self._json_ok()
        elif self.path == '/api/stop/add':
            name = str(data.get('name', '')).strip()
            if name and name not in STOP_LIST:
                STOP_LIST.append(name)
                stop_counters[name] = {'enter': 0, 'exit': 0}
            self._json_ok()
        elif self.path == '/api/stop/rename':
            idx = int(data.get('index', -1))
            name = str(data.get('name', '')).strip()
            if 0 <= idx < len(STOP_LIST) and name:
                old_name = STOP_LIST[idx]
                STOP_LIST[idx] = name
                stop_counters[name] = stop_counters.pop(old_name, {'enter': 0, 'exit': 0})
                if current_stop_index == idx:
                    STOP_NAME = name
            self._json_ok()
        elif self.path == '/api/stop/delete':
            idx = int(data.get('index', -1))
            if 0 <= idx < len(STOP_LIST) and len(STOP_LIST) > 1:
                old_name = STOP_LIST.pop(idx)
                stop_counters.pop(old_name, None)
                if current_stop_index >= len(STOP_LIST):
                    switch_to_stop(len(STOP_LIST) - 1)
                elif current_stop_index == idx:
                    STOP_NAME = STOP_LIST[current_stop_index]
                    load_stop_counters(STOP_NAME)
            self._json_ok()
        else:
            self.send_response(404)
            self.end_headers()

    def _json_ok(self):
        self.send_response(200)
        self.send_header('Content-Type', 'application/json')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        self.wfile.write(b'{"ok":true}')

    def log_message(self, format, *args):
        pass


def start_dashboard():
    server = HTTPServer(('0.0.0.0', DASHBOARD_PORT), DashboardHandler)
    server.daemon_threads = True
    t = threading.Thread(target=server.serve_forever, daemon=True)
    t.start()
    print(f'Веб-панель: http://localhost:{DASHBOARD_PORT}')


# ─── Многопоточное чтение камеры (FPS boost) ────────────────────────────────
class ThreadedVideoReader:
    """Читает кадры в фоновом потоке для минимальной задержки."""
    def __init__(self, cap):
        self.cap = cap
        self._frame = None
        self._ret = False
        self._lock = threading.Lock()
        self._running = True
        self._thread = threading.Thread(target=self._reader, daemon=True)
        self._thread.start()

    def _reader(self):
        while self._running:
            ret, frame = self.cap.read()
            if not ret or frame is None:
                # Не перетираем последний валидный кадр при кратковременном сбое чтения.
                time.sleep(0.01)
                continue
            with self._lock:
                self._ret = ret
                self._frame = frame

    def read(self):
        with self._lock:
            if self._frame is not None:
                return self._ret, self._frame.copy()
            return False, None

    def release(self):
        self._running = False
        self._thread.join(timeout=2)
        self.cap.release()

    def isOpened(self):
        return self.cap.isOpened()

    def set(self, prop, value):
        return self.cap.set(prop, value)

    def get(self, prop):
        return self.cap.get(prop)


def _font_path_preferred():
    candidates = [
        r'C:\Windows\Fonts\segoeui.ttf',
        r'C:\Windows\Fonts\arial.ttf',
        r'C:\Windows\Fonts\tahoma.ttf',
    ]
    for p in candidates:
        if os.path.exists(p):
            return p
    return None


def draw_text_unicode(img_bgr, text, x, y, font_px=26, color=(255, 255, 255), center=False):
    """Рисует UTF-8 текст (включая кириллицу) на BGR-кадре."""
    if Image is None or ImageDraw is None or ImageFont is None:
        # Fallback без кириллицы
        cv2.putText(img_bgr, str(text), (x, y), cv2.FONT_HERSHEY_SIMPLEX, 0.7, color, 1, cv2.LINE_AA)
        return

    rgb = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2RGB)
    pil_img = Image.fromarray(rgb)
    draw = ImageDraw.Draw(pil_img)

    font = _get_cached_font(font_px)

    tx, ty = x, y
    if center:
        bbox = draw.textbbox((0, 0), str(text), font=font)
        tw = bbox[2] - bbox[0]
        tx = max(0, int((img_bgr.shape[1] - tw) / 2))

    draw.text((tx, ty), str(text), font=font, fill=(int(color[2]), int(color[1]), int(color[0])))
    out = cv2.cvtColor(np.array(pil_img), cv2.COLOR_RGB2BGR)
    img_bgr[:, :] = out


_font_cache = {}
_panel_bg_cache = {}


def _get_cached_font(font_px):
    if ImageFont is None:
        return None
    fp = _font_path_preferred()
    cache_key = (fp or 'default', int(font_px))
    if cache_key not in _font_cache:
        try:
            _font_cache[cache_key] = ImageFont.truetype(fp, int(font_px)) if fp else ImageFont.load_default()
        except Exception:
            _font_cache[cache_key] = ImageFont.load_default()
    return _font_cache[cache_key]


def draw_text_unicode_batch(img_bgr, items):
    if not items:
        return
    if Image is None or ImageDraw is None or ImageFont is None:
        for item in items:
            cv2.putText(
                img_bgr,
                str(item['text']),
                (int(item['x']), int(item['y']) + int(item['font_px'])),
                cv2.FONT_HERSHEY_SIMPLEX,
                max(0.4, item['font_px'] / 42.0),
                item['color'],
                1,
                cv2.LINE_AA,
            )
        return

    rgb = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2RGB)
    pil_img = Image.fromarray(rgb)
    draw = ImageDraw.Draw(pil_img)

    for item in items:
        font = _get_cached_font(item['font_px'])
        tx, ty = int(item['x']), int(item['y'])
        if item.get('center'):
            bbox = draw.textbbox((0, 0), str(item['text']), font=font)
            tw = bbox[2] - bbox[0]
            tx = max(0, int((img_bgr.shape[1] - tw) / 2))
        draw.text(
            (tx, ty),
            str(item['text']),
            font=font,
            fill=(int(item['color'][2]), int(item['color'][1]), int(item['color'][0])),
        )

    out = cv2.cvtColor(np.array(pil_img), cv2.COLOR_RGB2BGR)
    img_bgr[:, :] = out


def get_panel_background(width, height):
    cache_key = (int(width), int(height))
    if cache_key in _panel_bg_cache:
        return _panel_bg_cache[cache_key]

    panel = np.zeros((height, width, 3), dtype=np.uint8)
    panel[:] = (24, 18, 10)
    cv2.line(panel, (0, 0), (width, 0), (40, 180, 255), 3)
    cv2.line(panel, (0, 1), (width, 1), (20, 110, 170), 1)
    _panel_bg_cache[cache_key] = panel
    return panel


def boxes_overlap_or_close(b1, b2, max_gap=60):
    x1, y1, w1, h1 = b1
    x2, y2, w2, h2 = b2

    if x1 <= x2 + w2 + max_gap and x1 + w1 + max_gap >= x2 and y1 <= y2 + h2 + max_gap and y1 + h1 + max_gap >= y2:
        return True
    return False


def merge_boxes(boxes, max_gap=80):
    """Объединяет близкие или перекрывающиеся боксы (руки + тело в один)"""
    merged = boxes.copy()
    i = 0
    while i < len(merged):
        j = i + 1
        while j < len(merged):
            if boxes_overlap_or_close(merged[i], merged[j], max_gap):
                x1, y1, w1, h1 = merged[i]
                x2, y2, w2, h2 = merged[j]
                nx = min(x1, x2)
                ny = min(y1, y2)
                nw = max(x1 + w1, x2 + w2) - nx
                nh = max(y1 + h1, y2 + h2) - ny
                merged[i] = (nx, ny, nw, nh)
                merged.pop(j)
                j = i + 1
            else:
                j += 1
        i += 1
    return merged


def init_database():
    if not os.path.exists(REPORT_FOLDER):
        os.makedirs(REPORT_FOLDER, exist_ok=True)

    conn = sqlite3.connect(DATABASE_PATH)
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS passenger_flow (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT,
            route TEXT,
            vehicle TEXT,
            stop TEXT,
            door INTEGER,
            direction TEXT,
            event_type TEXT
        )
    ''')
    conn.commit()
    conn.close()


def log_event(direction, event_type='crossing'):
    ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    conn = sqlite3.connect(DATABASE_PATH)
    c = conn.cursor()
    c.execute('''
        INSERT INTO passenger_flow (timestamp, route, vehicle, stop, door, direction, event_type)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (ts, ROUTE_NAME, VEHICLE_NAME, STOP_NAME, DOOR_NUMBER, direction, event_type))
    conn.commit()
    conn.close()


def show_splash():
    """Заставка НМУ ВКС при запуске — Программа 2: Определение траекторий."""
    W, H = 900, 500
    img = np.zeros((H, W, 3), dtype=np.uint8)
    img[:] = (60, 30, 10)  # тёмно-синий фон

    cv2.rectangle(img, (0, 0), (W, 8), (0, 180, 255), -1)
    cv2.rectangle(img, (0, H - 8), (W, H), (0, 180, 255), -1)

    draw_text_unicode(img, ORG_NAME, 0, 60, font_px=74, color=(0, 200, 255), center=True)
    draw_text_unicode(img, ORG_CITY, 0, 140, font_px=32, color=(180, 180, 180), center=True)
    # Пожелание заказчика: новое название программы 2
    draw_text_unicode(img, 'Определение траекторий', 0, 200, font_px=46, color=(100, 255, 150), center=True)
    draw_text_unicode(img, 'движения пассажиров', 0, 260, font_px=46, color=(100, 255, 150), center=True)
    draw_text_unicode(img, 'MediaPipe  |  OpenCV  |  YOLOv8  |  SQLite  |  Excel', 0, 320, font_px=22, color=(150, 150, 150), center=True)
    draw_text_unicode(img, f'{ORG_PHONE}  |  {ORG_EMAIL}', 0, 378, font_px=34, color=(100, 200, 100), center=True)
    draw_text_unicode(img, 'Нажмите любую клавишу для запуска...', 0, 440, font_px=26, color=(200, 200, 200), center=True)

    cv2.imshow('People Counter Splash', img)
    cv2.waitKey(0)
    cv2.destroyWindow('People Counter Splash')


def generate_reports():
    conn = sqlite3.connect(DATABASE_PATH)
    df = pd.read_sql_query('SELECT * FROM passenger_flow', conn)
    conn.close()

    if df.empty:
        print('Нет данных для отчёта.')
        return

    df['timestamp'] = pd.to_datetime(df['timestamp'])
    df['date']  = df['timestamp'].dt.date
    df['hour']  = df['timestamp'].dt.hour
    df['month'] = df['timestamp'].dt.month
    df['year']  = df['timestamp'].dt.year

    ctrl = df.groupby(['route', 'vehicle', 'door', 'direction']).agg(total=('id', 'count')).reset_index()

    reports = {
        'daily':     df.groupby('date').size().reset_index(name='count'),
        'hourly':    df.groupby(['date', 'hour']).size().reset_index(name='count'),
        'stop':      df.groupby('stop').size().reset_index(name='count'),
        'vehicle':   df.groupby('vehicle').size().reset_index(name='count'),
        'monthly':   df.groupby(['year', 'month']).size().reset_index(name='count'),
        'yearly':    df.groupby('year').size().reset_index(name='count'),
        'route':     df.groupby(['route', 'door']).size().reset_index(name='count'),
        'direction': df.groupby(['direction']).size().reset_index(name='count'),
    }

    for key, table in reports.items():
        table.to_csv(os.path.join(REPORT_FOLDER, f'{key}_report.csv'), index=False, encoding='utf-8-sig')

    ctrl.to_csv(os.path.join(REPORT_FOLDER, 'control_panel_report.csv'), index=False, encoding='utf-8-sig')

    # Матрица корреспонденций: сколько вошло/вышло на каждой остановке
    entered = df[df['direction'] == 'in'].groupby('stop').size().rename('вошло')
    exited  = df[df['direction'] == 'out'].groupby('stop').size().rename('вышло')
    matrix  = pd.concat([entered, exited], axis=1).fillna(0).astype(int)
    matrix['баланс'] = matrix['вошло'] - matrix['вышло']
    matrix.index.name = 'остановка'
    matrix.to_csv(os.path.join(REPORT_FOLDER, 'correspondence_matrix.csv'), encoding='utf-8-sig')

    # ─── АНАЛИЗ ТРАНЗИТНЫХ ПАССАЖИРОВ (Пожелание заказчика) ──────────────────
    # Пассажиры, вошедшие и вышедшие на ОДНОЙ остановке за время ≤ TIME_AT_STOP_SEC
    transit_df = None
    if OD_AVAILABLE and OD_TRACKING_ENABLED:
        try:
            conn_tmp = sqlite3.connect(DATABASE_PATH)
            # Загружаем OD-логи с временной информацией
            od_logs = pd.read_sql_query('''
                SELECT id, passenger_id, from_stop, to_stop, entry_time, exit_time, time_on_board_sec
                FROM od_log
                WHERE from_stop = to_stop AND to_stop IS NOT NULL
            ''', conn_tmp)
            conn_tmp.close()

            if not od_logs.empty:
                # Преобразуем времена и вычисляем длительность
                od_logs['entry_time'] = pd.to_datetime(od_logs['entry_time'], errors='coerce')
                od_logs['exit_time'] = pd.to_datetime(od_logs['exit_time'], errors='coerce')
                od_logs['duration_sec'] = od_logs['time_on_board_sec'].fillna(0)

                # Фильтруем транзиты: вошли и вышли на одной остановке за TIME_AT_STOP_SEC
                transits = od_logs[
                    (od_logs['from_stop'] == od_logs['to_stop']) &
                    (od_logs['duration_sec'] <= TIME_AT_STOP_SEC) &
                    (od_logs['duration_sec'] > 0)
                ].copy()

                if not transits.empty:
                    transit_df = transits[[
                        'passenger_id', 'from_stop', 'entry_time', 'exit_time', 'duration_sec'
                    ]].rename(columns={
                        'passenger_id': 'Пассажир',
                        'from_stop': 'Остановка',
                        'entry_time': 'Время входа',
                        'exit_time': 'Время выхода',
                        'duration_sec': 'Время в салоне (сек)',
                    })
                    transit_df.to_csv(
                        os.path.join(REPORT_FOLDER, 'transit_passengers_report.csv'),
                        index=False, encoding='utf-8-sig'
                    )
                    print(f'Отчёт транзитных пассажиров: {len(transits)} записей (≤{TIME_AT_STOP_SEC}сек)')
        except Exception as e:
            print(f'Анализ транзитов: {e}')

    # Отчёты по траекториям и внешности пассажиров
    traj_df = None
    app_df = None
    if IDENT_AVAILABLE and PASSENGER_ID_ENABLED:
        try:
            _tmp_ident = PassengerIdentifier()
            _tmp_db = PassengerDB(DATABASE_PATH, _tmp_ident, PASSENGER_ID_THRESHOLD)
            # Траектории: путь каждого пассажира по остановкам
            trajectories = _tmp_db.get_all_trajectories()
            if trajectories:
                traj_df = pd.DataFrame(trajectories)
                traj_df.to_csv(os.path.join(REPORT_FOLDER, 'trajectory_report.csv'),
                               index=False, encoding='utf-8-sig')
                print(f'Отчёт траекторий: {len(trajectories)} записей')
            # Параметры внешности: рост, одежда, цвета
            appearances = _tmp_db.get_all_appearances()
            if appearances:
                app_df = pd.DataFrame(appearances)
                app_df.to_csv(os.path.join(REPORT_FOLDER, 'appearance_report.csv'),
                              index=False, encoding='utf-8-sig')
                print(f'Отчёт внешности: {len(appearances)} записей')
            _tmp_ident.close()
        except Exception as e:
            print(f'Ошибка отчётов идентификации: {e}')

    try:
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        xlsx_path = os.path.join(REPORT_FOLDER, 'passenger_flow_report.xlsx')
        writer = pd.ExcelWriter(xlsx_path, engine='openpyxl')

        header_colors = {
            'Все данные':          '1F4E79',
            'По дням':             '2E75B6',
            'По часам':            '2E75B6',
            'По месяцам':          '375623',
            'По годам':            '375623',
            'По остановкам':       '7B2C2C',
            'По маршрутам':        '4472C4',
            'Панель руководителя': '833C00',
            'Матрица корреспонд.': '4C4C4C',
            'Траектории пассажиров': '1A5276',
            'Внешность пассажиров': '6C3483',
        }
        row_fill_odd  = PatternFill('solid', fgColor='DCE6F1')
        row_fill_even = PatternFill('solid', fgColor='FFFFFF')
        thin = Side(border_style='thin', color='CCCCCC')
        cell_border = Border(top=thin, left=thin, right=thin, bottom=thin)

        sheets_xlsx = {
            'Все данные': df[['id','timestamp','route','vehicle','stop','door','direction','event_type']].rename(columns={
                'id':'ID','timestamp':'Дата/Время','route':'Маршрут','vehicle':'ТС',
                'stop':'Остановка','door':'Дверь','direction':'Направление','event_type':'Событие'}),
            'По дням':    reports['daily'].rename(columns={'date':'Дата','count':'Пассажиров'}),
            'По часам':   reports['hourly'].rename(columns={'date':'Дата','hour':'Час','count':'Пассажиров'}),
            'По месяцам': reports['monthly'].rename(columns={'year':'Год','month':'Месяц','count':'Пассажиров'}),
            'По годам':   reports['yearly'].rename(columns={'year':'Год','count':'Пассажиров'}),
            'По остановкам': reports['stop'].rename(columns={'stop':'Остановка','count':'Пассажиров'}),
            'По маршрутам':  reports['route'].rename(columns={'route':'Маршрут','door':'Дверь','count':'Пассажиров'}),
            'Панель руководителя': ctrl.rename(columns={'route':'Маршрут','vehicle':'ТС','door':'Дверь','direction':'Напр.','total':'Итого'}),
            'Матрица корреспонд.': matrix.reset_index().rename(columns={'остановка':'Остановка'}),
        }

        # Добавляем листы траекторий и внешности, если данные есть
        if traj_df is not None and not traj_df.empty:
            sheets_xlsx['Траектории пассажиров'] = traj_df.rename(columns={
                'passenger_id': 'Пассажир',
                'route': 'Маршрут',
                'entry_stop': 'Ост. входа',
                'exit_stop': 'Ост. выхода',
                'stops_visited': 'Траектория',
                'stops_count': 'Кол-во ост.',
                'entry_time': 'Время входа',
                'exit_time': 'Время выхода',
            })
        if app_df is not None and not app_df.empty:
            sheets_xlsx['Внешность пассажиров'] = app_df.rename(columns={
                'passenger_id': 'Пассажир',
                'timestamp': 'Время',
                'estimated_height_cm': 'Рост (см)',
                'upper_color_name': 'Цвет верха',
                'lower_color_name': 'Цвет низа',
                'hair_color_name': 'Цвет волос',
                'clothing_type': 'Тип одежды',
                'upper_body_color': 'RGB верх',
                'lower_body_color': 'RGB низ',
                'hair_color': 'RGB волос',
            })
        
        # Лист транзитных пассажиров (пожелание заказчика)
        if transit_df is not None and not transit_df.empty:
            sheets_xlsx['Транзитные пассажиры'] = transit_df
            header_colors['Транзитные пассажиры'] = '8B4513'

        # ── День 10: расширенные отчёты ─────────────────────────────────
        # 10-а. Лист «Сшивка маршрутов» — цепочки поездок из cross_route_stitcher
        try:
            from cross_route_stitcher import (
                load_passengers, load_trajectories,
                match_passengers_across_dbs, stitch_trajectories
            )
            _pax = load_passengers(DATABASE_PATH)
            _traj = load_trajectories(DATABASE_PATH)
            if _pax:
                _mapping = match_passengers_across_dbs(_pax)
                if _traj:
                    _chains = stitch_trajectories(_traj, _mapping)
                    if _chains:
                        chains_df = pd.DataFrame(_chains)
                        sheets_xlsx['Сшивка маршрутов'] = chains_df.rename(columns={
                            'global_id': 'Глобальный ID',
                            'legs': 'Рейсов',
                            'routes': 'Маршруты',
                            'origin': 'Начало',
                            'destination': 'Конец',
                            'full_path': 'Полный путь',
                            'start_time': 'Время начала',
                            'end_time': 'Время конца',
                            'duration_min': 'Длительность (мин)',
                            'transfers': 'Пересадок',
                        })
                        header_colors['Сшивка маршрутов'] = '2C3E50'
                        print(f'Отчёт сшивки: {len(_chains)} цепочек')
        except Exception as e:
            print(f'Сшивка маршрутов для Excel: {e}')

        # 10-б. Лист «OD сводка» — расширенная матрица с итогами и временем поездок
        if OD_AVAILABLE:
            try:
                _od = ODMatrixTracker(DATABASE_PATH)
                _od_matrix = _od.get_od_matrix()
                if _od_matrix:
                    _od_rows = []
                    for fr, tos in _od_matrix.items():
                        for to, cnt in tos.items():
                            _od_rows.append({'Откуда': fr, 'Куда': to, 'Пассажиров': cnt})
                    _od_df = pd.DataFrame(_od_rows)
                    # Добавляем строку итогов
                    _total_row = pd.DataFrame([{
                        'Откуда': 'ИТОГО', 'Куда': '',
                        'Пассажиров': int(_od_df['Пассажиров'].sum())
                    }])
                    _od_df = pd.concat([_od_df, _total_row], ignore_index=True)
                    sheets_xlsx['OD сводка'] = _od_df
                    header_colors['OD сводка'] = '1B4F72'
                # Статистика времени в пути
                _dur = _od.get_duration_stats()
                if _dur['total_trips'] > 0:
                    _dur_df = pd.DataFrame([{
                        'Показатель': 'Среднее время (сек)',
                        'Значение': round(_dur['avg_sec'], 1),
                    }, {
                        'Показатель': 'Мин. время (сек)',
                        'Значение': round(_dur['min_sec'], 1) if _dur['min_sec'] else '-',
                    }, {
                        'Показатель': 'Макс. время (сек)',
                        'Значение': round(_dur['max_sec'], 1) if _dur['max_sec'] else '-',
                    }, {
                        'Показатель': 'Всего поездок',
                        'Значение': _dur['total_trips'],
                    }])
                    sheets_xlsx['Время в пути'] = _dur_df
                    header_colors['Время в пути'] = '1A5276'
            except Exception as e:
                print(f'OD сводка для Excel: {e}')

        for sheet_name, data in sheets_xlsx.items():
            data.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
            ws = writer.sheets[sheet_name]
            color = header_colors.get(sheet_name, '1F4E79')

            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(data.columns))
            tc = ws.cell(row=1, column=1)
            tc.value = sheet_name
            tc.font  = Font(bold=True, size=13, color='FFFFFF')
            tc.fill  = PatternFill('solid', fgColor=color)
            tc.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 26

            for cell in ws[2]:
                cell.font  = Font(bold=True, color='FFFFFF')
                cell.fill  = PatternFill('solid', fgColor=color)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = cell_border
            ws.row_dimensions[2].height = 20

            for ri, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row)):
                fill = row_fill_odd if ri % 2 == 0 else row_fill_even
                for cell in row:
                    cell.fill = fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = cell_border

            for ci, col in enumerate(ws.columns, start=1):
                ml = max((len(str(c.value)) for c in col if c.value), default=8)
                ws.column_dimensions[get_column_letter(ci)].width = min(ml + 4, 45)

            ws.freeze_panes = 'A3'

        writer.close()
        print(f'Excel-отчёт: {xlsx_path}')
    except Exception as e:
        print('Ошибка Excel:', e)

    print('Все отчёты сформированы.')

# Инициализация счётчиков
enter_count = 0
exit_count = 0

# OD-трекер
od_tracker = None
if OD_AVAILABLE and OD_TRACKING_ENABLED:
    try:
        od_tracker = ODMatrixTracker(DATABASE_PATH)
        print('OD-матрица: активна')
    except Exception as _e_od:
        print(f'OD-матрица: ошибка инициализации ({_e_od})')

# Словарь для отслеживания объектов: {id: (prev_center, crossed, last_cross_direction)}
tracked_objects = {}
next_id = 0


def mouse_callback(event, x, y, flags, param):
    global LINE_ORIENTATION, LINE_Y, LINE_X, LINE_START_X, LINE_END_X, LINE_START_Y, LINE_END_Y, draw_state
    global current_stop_index, STOP_NAME, frame_width, frame_height, settings_request

    # Клик по панели (ниже кадра) — проверяем кнопки
    if event == cv2.EVENT_LBUTTONDOWN and frame_height > 0 and y >= frame_height:
        py = y - frame_height  # Y внутри панели
        for btn_name, (bx1, by1, bx2, by2) in panel_buttons.items():
            if bx1 <= x <= bx2 and by1 <= py <= by2:
                if btn_name == 'prev' and current_stop_index > 0:
                    switch_to_stop(current_stop_index - 1)
                elif btn_name == 'next' and current_stop_index < len(STOP_LIST) - 1:
                    switch_to_stop(current_stop_index + 1)
                elif btn_name == 'manage':
                    manage_stops_gui()
                elif btn_name == 'settings':
                    settings_request = True
                return  # не рисуем линию
        return  # клик по панели, но не по кнопке

    if event == cv2.EVENT_LBUTTONDOWN:
        draw_state['drawing'] = True
        draw_state['start'] = (x, y)
        draw_state['end'] = (x, y)

    elif event == cv2.EVENT_MOUSEMOVE and draw_state['drawing']:
        draw_state['end'] = (x, y)

    elif event == cv2.EVENT_LBUTTONUP and draw_state['drawing']:
        draw_state['drawing'] = False
        draw_state['end'] = (x, y)

        sx, sy = draw_state['start']
        ex, ey = draw_state['end']

        if max(abs(ex - sx), abs(ey - sy)) <= LINE_CLICK_MOVE_THRESHOLD:
            if LINE_ORIENTATION == 'horizontal':
                LINE_Y = min(max(y, 0), max(0, frame_height - 1))
                LINE_START_X, LINE_END_X = 0, max(1, frame_width - 1)
            else:
                LINE_X = min(max(x, 0), max(0, frame_width - 1))
                LINE_START_Y, LINE_END_Y = 0, max(1, frame_height - 1)
            return

        # Определяем ориентацию линии по большей величине проекции
        if abs(ex - sx) >= abs(ey - sy):
            LINE_ORIENTATION = 'horizontal'
            LINE_Y = int((sy + ey) / 2)
            LINE_START_X, LINE_END_X = min(sx, ex), max(sx, ex)

        else:
            LINE_ORIENTATION = 'vertical'
            LINE_X = int((sx + ex) / 2)
            LINE_START_Y, LINE_END_Y = min(sy, ey), max(sy, ey)


def open_video_source(source):
    """Открывает камеру/URL/файл с fallback-стратегией для телефонных потоков."""
    def source_kind(src):
        if isinstance(src, int):
            return 'usb'
        if isinstance(src, str):
            s = src.strip().lower()
            if s.startswith('rtsp://'):
                return 'rtsp'
            if s.startswith('http://') or s.startswith('https://'):
                return 'http'
            return 'file'
        return 'other'

    candidates = []

    if isinstance(source, int):
        candidates = [source]
    elif isinstance(source, str):
        s = source.strip()
        if s.startswith(('http://', 'https://')):
            parsed = urlparse(s)
            host = parsed.hostname or PHONE_IP
            port = parsed.port or PHONE_PORT
            scheme = parsed.scheme or 'http'
            base = f'{scheme}://{host}:{port}'
            candidates.append(s)
            for path in PHONE_FALLBACK_PATHS:
                url = f'{base}{path}'
                if url not in candidates:
                    candidates.append(url)
        else:
            candidates = [s]
    else:
        candidates = [source]

    for cand in candidates:
        kind = source_kind(cand)
        if kind in ('http', 'rtsp'):
            backends = []
            if hasattr(cv2, 'CAP_FFMPEG'):
                backends.append(cv2.CAP_FFMPEG)
            if hasattr(cv2, 'CAP_GSTREAMER'):
                backends.append(cv2.CAP_GSTREAMER)
            backends.append(None)
        elif kind == 'usb':
            backends = []
            if hasattr(cv2, 'CAP_DSHOW'):
                backends.append(cv2.CAP_DSHOW)
            if hasattr(cv2, 'CAP_MSMF'):
                backends.append(cv2.CAP_MSMF)
            backends.append(None)
        else:
            backends = [None]
            if hasattr(cv2, 'CAP_FFMPEG'):
                backends.append(cv2.CAP_FFMPEG)

        # Удаляем дубликаты backend-ов, сохраняя порядок.
        dedup_backends = []
        seen = set()
        for b in backends:
            key = 'none' if b is None else int(b)
            if key not in seen:
                seen.add(key)
                dedup_backends.append(b)

        for backend in dedup_backends:
            cap = cv2.VideoCapture(cand) if backend is None else cv2.VideoCapture(cand, backend)
            if not cap.isOpened():
                cap.release()
                continue

            # Для сетевых потоков уменьшаем зависания при подключении/чтении.
            if kind in ('http', 'rtsp'):
                if hasattr(cv2, 'CAP_PROP_OPEN_TIMEOUT_MSEC'):
                    cap.set(cv2.CAP_PROP_OPEN_TIMEOUT_MSEC, 4000)
                if hasattr(cv2, 'CAP_PROP_READ_TIMEOUT_MSEC'):
                    cap.set(cv2.CAP_PROP_READ_TIMEOUT_MSEC, 2000)
                cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)

            # Несколько попыток чтения первого кадра.
            max_tries = 16 if kind in ('http', 'rtsp') else 6
            ok = False
            for _ in range(max_tries):
                ret, _ = cap.read()
                if ret:
                    ok = True
                    break
                time.sleep(0.08)
            if ok:
                backend_name = 'default' if backend is None else str(backend)
                print(f'Источник открыт: {cand} | backend={backend_name}')
                return cap

            cap.release()

    return cv2.VideoCapture()  # пустой объект с isOpened() == False


def get_screen_size():
    """Возвращает размер экрана (пиксели)."""
    try:
        import tkinter as tk
        root = tk.Tk()
        root.withdraw()
        w = int(root.winfo_screenwidth())
        h = int(root.winfo_screenheight())
        root.destroy()
        return w, h
    except Exception:
        return 1920, 1080


def fit_frame_to_screen(base_w, base_h, panel_h):
    """Подгоняет размер кадра так, чтобы окно (кадр + панель) всегда помещалось на экран."""
    if not AUTO_FIT_TO_SCREEN:
        return int(base_w), int(base_h)

    scr_w, scr_h = get_screen_size()
    max_w = max(320, scr_w - WINDOW_MARGIN_W)
    max_h = max(240, scr_h - panel_h - WINDOW_MARGIN_H)

    scale = min(max_w / float(base_w), max_h / float(base_h), 1.0)
    out_w = max(320, int(base_w * scale))
    out_h = max(240, int(base_h * scale))

    # На всякий случай никогда не выходим за доступный экран.
    out_w = min(out_w, max_w)
    out_h = min(out_h, max_h)

    # Чётные размеры стабильнее для некоторых кодеков/бэкендов.
    out_w -= out_w % 2
    out_h -= out_h % 2
    return out_w, out_h


def key_is(key, *values):
    if key < 0:
        return False
    low = key & 0xFF
    for v in values:
        if isinstance(v, int) and key == v:
            return True
        if isinstance(v, str) and len(v) == 1:
            code = ord(v)
            if key == code or low == code:
                return True
    return False


def fit_dialog_geometry(root, target_w, target_h, margin_w=40, margin_h=80):
    """Подгоняет окно настроек под экран, чтобы кнопки не уезжали за пределы."""
    try:
        scr_w, scr_h = get_screen_size()
        w = min(int(target_w), max(420, scr_w - margin_w))
        h = min(int(target_h), max(360, scr_h - margin_h))
        x = max(0, (scr_w - w) // 2)
        y = max(0, (scr_h - h) // 2)
        root.geometry(f'{w}x{h}+{x}+{y}')
    except Exception:
        root.geometry(f'{target_w}x{target_h}')


def apply_capture_profile(cap, source):
    """Применяет безопасные параметры захвата для типа источника."""
    cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)  # минимальный буфер — всегда свежий кадр

    is_usb_cam = isinstance(source, int)
    if is_usb_cam:
        # Для USB-камер: безопасный профиль чтобы избежать мигания/полос.
        cap.set(cv2.CAP_PROP_FPS, USB_CAM_TARGET_FPS)
        if not USB_CAM_SAFE_MODE:
            cap.set(cv2.CAP_PROP_FOURCC, cv2.VideoWriter_fourcc(*'MJPG'))
            print('USB-камера: производительный профиль (MJPG).')
        else:
            print('USB-камера: безопасный профиль (без принудительного MJPG).')
    # Для RTSP/HTTP/файла FOURCC не форсируем.

    if isinstance(source, str) and source.lower().startswith(('rtsp://', 'http://', 'https://')):
        if hasattr(cv2, 'CAP_PROP_OPEN_TIMEOUT_MSEC'):
            cap.set(cv2.CAP_PROP_OPEN_TIMEOUT_MSEC, 4000)
        if hasattr(cv2, 'CAP_PROP_READ_TIMEOUT_MSEC'):
            cap.set(cv2.CAP_PROP_READ_TIMEOUT_MSEC, 2000)


def manage_stops_gui():
    """Окно управления остановками: добавить, удалить, переименовать, переместить."""
    global STOP_LIST, current_stop_index, STOP_NAME
    import tkinter as tk
    from tkinter import simpledialog, messagebox

    win = tk.Tk()
    win.title('Управление остановками')
    fit_dialog_geometry(win, 420, 400)
    win.minsize(420, 360)
    win.resizable(True, True)

    tk.Label(win, text='Остановки маршрута', font=('Segoe UI', 12, 'bold')).pack(pady=(10, 4))

    frm = tk.Frame(win)
    frm.pack(fill='both', expand=True, padx=12)

    listbox = tk.Listbox(frm, font=('Segoe UI', 11), selectmode='single', height=12)
    listbox.pack(side='left', fill='both', expand=True)
    scrollbar = tk.Scrollbar(frm, command=listbox.yview)
    scrollbar.pack(side='right', fill='y')
    listbox.config(yscrollcommand=scrollbar.set)

    def refresh_list():
        listbox.delete(0, tk.END)
        for i, s in enumerate(STOP_LIST):
            prefix = '▶ ' if i == current_stop_index else '   '
            listbox.insert(tk.END, f'{prefix}{i+1}. {s}')
        if current_stop_index < listbox.size():
            listbox.selection_set(current_stop_index)

    refresh_list()

    btn_frame = tk.Frame(win)
    btn_frame.pack(fill='x', padx=12, pady=8)

    def add_stop():
        name = simpledialog.askstring('Добавить', 'Название новой остановки:', parent=win)
        if name and name.strip():
            STOP_LIST.append(name.strip())
            rebuild_stop_counters()
            refresh_list()

    def rename_stop():
        global STOP_NAME
        sel = listbox.curselection()
        if not sel:
            return
        idx = sel[0]
        old = STOP_LIST[idx]
        name = simpledialog.askstring('Переименовать', f'Новое название для «{old}»:', initialvalue=old, parent=win)
        if name and name.strip():
            new_name = name.strip()
            old_counter = stop_counters.get(old, {'enter': 0, 'exit': 0})
            STOP_LIST[idx] = new_name
            stop_counters.pop(old, None)
            stop_counters[new_name] = old_counter
            if current_stop_index == idx:
                STOP_NAME = new_name
            rebuild_stop_counters()
            refresh_list()

    def delete_stop():
        global current_stop_index
        sel = listbox.curselection()
        if not sel:
            return
        idx = sel[0]
        if len(STOP_LIST) <= 1:
            messagebox.showwarning('Нельзя', 'Должна остаться хотя бы одна остановка.', parent=win)
            return
        STOP_LIST.pop(idx)
        if current_stop_index >= len(STOP_LIST):
            current_stop_index = len(STOP_LIST) - 1
        rebuild_stop_counters()
        refresh_list()

    def move_up():
        sel = listbox.curselection()
        if not sel or sel[0] == 0:
            return
        global current_stop_index
        idx = sel[0]
        STOP_LIST[idx], STOP_LIST[idx-1] = STOP_LIST[idx-1], STOP_LIST[idx]
        if current_stop_index == idx:
            current_stop_index -= 1
        elif current_stop_index == idx - 1:
            current_stop_index += 1
        refresh_list()
        listbox.selection_set(idx - 1)

    def move_down():
        sel = listbox.curselection()
        if not sel or sel[0] >= len(STOP_LIST) - 1:
            return
        global current_stop_index
        idx = sel[0]
        STOP_LIST[idx], STOP_LIST[idx+1] = STOP_LIST[idx+1], STOP_LIST[idx]
        if current_stop_index == idx:
            current_stop_index += 1
        elif current_stop_index == idx + 1:
            current_stop_index -= 1
        refresh_list()
        listbox.selection_set(idx + 1)

    def select_stop():
        global current_stop_index, STOP_NAME
        sel = listbox.curselection()
        if not sel:
            return
        current_stop_index = sel[0]
        STOP_NAME = STOP_LIST[current_stop_index]
        refresh_list()

    def close_win():
        global STOP_NAME
        STOP_NAME = STOP_LIST[current_stop_index]
        rebuild_stop_counters()
        save_config()
        win.destroy()

    tk.Button(btn_frame, text='➕ Добавить', command=add_stop, width=12).grid(row=0, column=0, padx=2, pady=2)
    tk.Button(btn_frame, text='✏ Переимен.', command=rename_stop, width=12).grid(row=0, column=1, padx=2, pady=2)
    tk.Button(btn_frame, text='❌ Удалить', command=delete_stop, width=12).grid(row=0, column=2, padx=2, pady=2)
    tk.Button(btn_frame, text='⬆ Вверх', command=move_up, width=12).grid(row=1, column=0, padx=2, pady=2)
    tk.Button(btn_frame, text='⬇ Вниз', command=move_down, width=12).grid(row=1, column=1, padx=2, pady=2)
    tk.Button(btn_frame, text='✅ Выбрать', command=select_stop, width=12).grid(row=1, column=2, padx=2, pady=2)

    tk.Button(win, text='Закрыть', command=close_win, width=16).pack(pady=(0, 10))
    win.protocol('WM_DELETE_WINDOW', close_win)
    win.grab_set()
    win.mainloop()


def choose_source_gui(default_source, allow_autostart=True,
                      window_title='Источник и параметры рейса',
                      action_text='Запустить'):
    """Окно настроек источника, маршрута и рабочих параметров."""
    import tkinter as tk
    from tkinter import filedialog, messagebox

    if allow_autostart and _config_loaded and AUTO_START_LAST_SOURCE and not FORCE_SOURCE_DIALOG and default_source not in (None, ''):
        print(f'Автозапуск с сохранённым источником: {default_source}')
        return default_source

    selected = {'value': default_source}

    root = tk.Tk()
    root.title(window_title)
    fit_dialog_geometry(root, 720, 660)
    root.minsize(640, 520)
    root.resizable(True, True)

    mode = tk.StringVar(value='phone')
    phone_ip = tk.StringVar(value=PHONE_IP)
    phone_port = tk.StringVar(value=str(PHONE_PORT))
    rtsp_url = tk.StringVar(value='rtsp://user:pass@192.168.0.10:554/stream')
    file_path = tk.StringVar(value='')
    route_name = tk.StringVar(value=ROUTE_NAME)
    vehicle_name = tk.StringVar(value=VEHICLE_NAME)
    door_number = tk.StringVar(value=str(DOOR_NUMBER))
    stop_list_text = tk.StringVar(value='; '.join(STOP_LIST))
    start_stop = tk.StringVar(value=str(current_stop_index + 1))
    count_point_mode = tk.StringVar(value=COUNT_POINT_MODE)
    reverse_direction = tk.BooleanVar(value=REVERSE_COUNT_DIRECTION)
    camera_rotation = tk.StringVar(value=str(CAMERA_ROTATION))
    yolo_confidence = tk.StringVar(value=f'{YOLO_CONFIDENCE:.2f}')
    passenger_id_threshold = tk.StringVar(value=f'{PASSENGER_ID_THRESHOLD:.2f}')
    yolo_every_frames = tk.StringVar(value=str(YOLO_EVERY_N_FRAMES))
    yolo_imgsz = tk.StringVar(value=str(YOLO_IMGSZ))
    min_area = tk.StringVar(value=str(MIN_AREA))
    track_dist_ratio = tk.StringVar(value=f'{TRACKING_MAX_DIST_RATIO:.2f}')
    track_lost_frames = tk.StringVar(value=str(TRACKING_LOST_FRAMES))
    time_at_stop = tk.StringVar(value=f'{TIME_AT_STOP_SEC:g}')

    if isinstance(default_source, int):
        mode.set('usb')
    elif isinstance(default_source, str) and default_source.startswith('http'):
        mode.set('phone')
    elif isinstance(default_source, str) and default_source.startswith('rtsp'):
        mode.set('rtsp')
        rtsp_url.set(default_source)
    elif isinstance(default_source, str):
        mode.set('file')
        file_path.set(default_source)

    tk.Label(root, text='Источник, маршрут и рабочие параметры', font=('Segoe UI', 12, 'bold')).pack(pady=10)

    frm = tk.Frame(root)
    frm.pack(fill='both', expand=True, padx=12)

    tk.Radiobutton(frm, text='USB камера (0)', variable=mode, value='usb').grid(row=0, column=0, sticky='w')
    tk.Radiobutton(frm, text='Телефон (IP Webcam)', variable=mode, value='phone').grid(row=1, column=0, sticky='w', pady=(8, 0))

    tk.Label(frm, text='IP').grid(row=2, column=0, sticky='w', padx=(20, 0))
    tk.Entry(frm, textvariable=phone_ip, width=18).grid(row=2, column=1, sticky='w')
    tk.Label(frm, text='Port').grid(row=2, column=2, sticky='w', padx=(10, 0))
    tk.Entry(frm, textvariable=phone_port, width=8).grid(row=2, column=3, sticky='w')

    tk.Radiobutton(frm, text='RTSP', variable=mode, value='rtsp').grid(row=3, column=0, sticky='w', pady=(10, 0))
    tk.Entry(frm, textvariable=rtsp_url, width=54).grid(row=4, column=0, columnspan=4, sticky='w', padx=(20, 0))

    tk.Radiobutton(frm, text='Видеофайл', variable=mode, value='file').grid(row=5, column=0, sticky='w', pady=(10, 0))
    tk.Entry(frm, textvariable=file_path, width=44).grid(row=6, column=0, columnspan=3, sticky='w', padx=(20, 0))

    def pick_file():
        path = filedialog.askopenfilename(
            title='Выберите видеофайл',
            filetypes=[('Video', '*.mp4 *.avi *.mov *.mkv'), ('All', '*.*')],
        )
        if path:
            file_path.set(path)
            mode.set('file')

    tk.Button(frm, text='Обзор...', command=pick_file).grid(row=6, column=3, sticky='w', padx=(8, 0))

    tk.Label(frm, text='Маршрут').grid(row=7, column=0, sticky='w', pady=(16, 0))
    tk.Entry(frm, textvariable=route_name, width=24).grid(row=7, column=1, sticky='w', pady=(16, 0))
    tk.Label(frm, text='Автобус').grid(row=7, column=2, sticky='w', pady=(16, 0))
    tk.Entry(frm, textvariable=vehicle_name, width=16).grid(row=7, column=3, sticky='w', pady=(16, 0))

    tk.Label(frm, text='Дверь').grid(row=8, column=0, sticky='w', pady=(8, 0))
    tk.Entry(frm, textvariable=door_number, width=8).grid(row=8, column=1, sticky='w', pady=(8, 0))
    tk.Label(frm, text='Старт. остановка №').grid(row=8, column=2, sticky='w', pady=(8, 0))
    tk.Entry(frm, textvariable=start_stop, width=8).grid(row=8, column=3, sticky='w', pady=(8, 0))

    tk.Label(frm, text='Остановки через ;').grid(row=9, column=0, sticky='nw', pady=(10, 0))
    tk.Entry(frm, textvariable=stop_list_text, width=54).grid(row=9, column=1, columnspan=3, sticky='w', pady=(10, 0))

    tk.Label(frm, text='Точка счёта').grid(row=10, column=0, sticky='w', pady=(10, 0))
    tk.Radiobutton(frm, text='Центр', variable=count_point_mode, value='center').grid(row=10, column=1, sticky='w', pady=(10, 0))
    tk.Radiobutton(frm, text='Нижняя точка', variable=count_point_mode, value='bottom').grid(row=10, column=2, sticky='w', pady=(10, 0))
    tk.Checkbutton(frm, text='Поменять вход/выход местами', variable=reverse_direction).grid(row=11, column=0, columnspan=4, sticky='w', pady=(8, 0))

    tk.Label(frm, text='YOLO conf').grid(row=12, column=0, sticky='w', pady=(14, 0))
    tk.Entry(frm, textvariable=yolo_confidence, width=10).grid(row=12, column=1, sticky='w', pady=(14, 0))
    tk.Label(frm, text='ID порог').grid(row=12, column=2, sticky='w', pady=(14, 0))
    tk.Entry(frm, textvariable=passenger_id_threshold, width=10).grid(row=12, column=3, sticky='w', pady=(14, 0))

    tk.Label(frm, text='YOLO кадр N').grid(row=13, column=0, sticky='w', pady=(8, 0))
    tk.Entry(frm, textvariable=yolo_every_frames, width=10).grid(row=13, column=1, sticky='w', pady=(8, 0))
    tk.Label(frm, text='YOLO imgsz').grid(row=13, column=2, sticky='w', pady=(8, 0))
    tk.Entry(frm, textvariable=yolo_imgsz, width=10).grid(row=13, column=3, sticky='w', pady=(8, 0))

    tk.Label(frm, text='Transit сек').grid(row=14, column=0, sticky='w', pady=(8, 0))
    tk.Entry(frm, textvariable=time_at_stop, width=10).grid(row=14, column=1, sticky='w', pady=(8, 0))
    tk.Label(frm, text='Min area').grid(row=14, column=2, sticky='w', pady=(8, 0))
    tk.Entry(frm, textvariable=min_area, width=10).grid(row=14, column=3, sticky='w', pady=(8, 0))

    tk.Label(frm, text='Track dist').grid(row=15, column=0, sticky='w', pady=(8, 0))
    tk.Entry(frm, textvariable=track_dist_ratio, width=10).grid(row=15, column=1, sticky='w', pady=(8, 0))
    tk.Label(frm, text='Track lost').grid(row=15, column=2, sticky='w', pady=(8, 0))
    tk.Entry(frm, textvariable=track_lost_frames, width=10).grid(row=15, column=3, sticky='w', pady=(8, 0))

    tk.Label(frm, text='Разворот камеры').grid(row=16, column=0, sticky='w', pady=(8, 0))
    tk.Radiobutton(frm, text='0°', variable=camera_rotation, value='0').grid(row=16, column=1, sticky='w', pady=(8, 0))
    tk.Radiobutton(frm, text='90°', variable=camera_rotation, value='90').grid(row=16, column=2, sticky='w', pady=(8, 0))
    tk.Radiobutton(frm, text='180°', variable=camera_rotation, value='180').grid(row=16, column=3, sticky='w', pady=(8, 0))
    tk.Radiobutton(frm, text='270°', variable=camera_rotation, value='270').grid(row=17, column=1, sticky='w', pady=(4, 0))

    tk.Label(frm, text='Для лестницы обычно: Min area 600-900, Track dist 0.14-0.18').grid(row=18, column=0, columnspan=4, sticky='w', pady=(8, 0))

    status = tk.StringVar(value='Готово к запуску')
    tk.Label(root, textvariable=status, fg='#4a4a4a').pack(anchor='w', padx=14, pady=(4, 0))

    def parse_float_value(raw_value, fallback, min_value, max_value):
        try:
            value = float(str(raw_value).replace(',', '.'))
        except Exception:
            value = float(fallback)
        return max(min_value, min(max_value, value))

    def parse_int_value(raw_value, fallback, min_value, max_value):
        try:
            value = int(float(str(raw_value).replace(',', '.')))
        except Exception:
            value = int(fallback)
        return max(min_value, min(max_value, value))

    def build_source_from_ui():
        m = mode.get()
        if m == 'usb':
            return 0
        if m == 'phone':
            ip = phone_ip.get().strip()
            port = phone_port.get().strip() or '8080'
            return f'http://{ip}:{port}/video'
        if m == 'rtsp':
            return rtsp_url.get().strip()
        return file_path.get().strip()

    def test_connection():
        source = build_source_from_ui()
        if not source and mode.get() != 'usb':
            messagebox.showwarning('Пустой источник', 'Заполните параметры источника.')
            return

        status.set('Проверка подключения...')
        root.update_idletasks()
        cap = open_video_source(source)
        ok = cap.isOpened()
        if ok:
            ret, _ = cap.read()
            ok = bool(ret)
        cap.release()

        if ok:
            status.set('Подключение успешно')
            messagebox.showinfo('Успех', 'Источник доступен, можно запускать.')
        else:
            status.set('Не удалось подключиться')
            messagebox.showerror('Ошибка', 'Источник недоступен. Проверьте IP/порт/URL.')

    def cancel_selection():
        root.destroy()

    def run_with_selected():
        global YOLO_CONFIDENCE, PASSENGER_ID_THRESHOLD, YOLO_EVERY_N_FRAMES, YOLO_IMGSZ, TIME_AT_STOP_SEC
        global MIN_AREA, TRACKING_MAX_DIST_RATIO, TRACKING_LOST_FRAMES
        global CAMERA_ROTATION

        src = build_source_from_ui()
        if not src and mode.get() != 'usb':
            messagebox.showwarning('Пустой источник', 'Заполните параметры источника.')
            return

        if src != default_source:
            status.set('Проверка нового источника...')
            root.update_idletasks()
            cap = open_video_source(src)
            ok = cap.isOpened()
            cap.release()
            if not ok:
                status.set('Не удалось подключиться')
                messagebox.showerror('Ошибка', 'Новый источник не открывается. Проверьте параметры или нажмите Проверить подключение.')
                return

        parsed_stops = parse_stop_list(stop_list_text.get())
        try:
            start_idx = max(0, min(int(start_stop.get() or '1') - 1, len(parsed_stops) - 1))
        except Exception:
            start_idx = 0

        yolo_conf_value = parse_float_value(yolo_confidence.get(), YOLO_CONFIDENCE, 0.10, 0.95)
        id_threshold_value = parse_float_value(passenger_id_threshold.get(), PASSENGER_ID_THRESHOLD, 0.05, 0.60)
        yolo_every_value = parse_int_value(yolo_every_frames.get(), YOLO_EVERY_N_FRAMES, 1, 8)
        yolo_imgsz_value = parse_int_value(yolo_imgsz.get(), YOLO_IMGSZ, 256, 1280)
        yolo_imgsz_value = max(256, min(1280, int(round(yolo_imgsz_value / 32.0) * 32)))
        min_area_value = parse_int_value(min_area.get(), MIN_AREA, 200, 5000)
        track_dist_value = parse_float_value(track_dist_ratio.get(), TRACKING_MAX_DIST_RATIO, 0.05, 0.30)
        track_lost_value = parse_int_value(track_lost_frames.get(), TRACKING_LOST_FRAMES, 2, 60)
        rotation_value = parse_int_value(camera_rotation.get(), CAMERA_ROTATION, 0, 270)
        if rotation_value not in (0, 90, 180, 270):
            rotation_value = 0
        time_at_stop_value = parse_float_value(time_at_stop.get(), TIME_AT_STOP_SEC, 5.0, 300.0)

        save_current_stop_counters()
        apply_transport_settings(
            route_name=route_name.get(),
            vehicle_name=vehicle_name.get(),
            door_number=door_number.get(),
            stop_list=';'.join(parsed_stops),
            start_stop_index=start_idx,
        )
        apply_counting_settings(
            point_mode=count_point_mode.get(),
            reverse_direction=reverse_direction.get(),
        )
        YOLO_CONFIDENCE = yolo_conf_value
        PASSENGER_ID_THRESHOLD = id_threshold_value
        YOLO_EVERY_N_FRAMES = yolo_every_value
        YOLO_IMGSZ = yolo_imgsz_value
        MIN_AREA = min_area_value
        TRACKING_MAX_DIST_RATIO = track_dist_value
        TRACKING_LOST_FRAMES = track_lost_value
        CAMERA_ROTATION = rotation_value
        TIME_AT_STOP_SEC = time_at_stop_value
        sync_source_settings_from_source(src)
        save_config()
        status.set('Настройки сохранены')
        selected['value'] = src
        root.destroy()

    btns = tk.Frame(root)
    btns.pack(fill='x', padx=12, pady=12)
    tk.Button(btns, text='Проверить подключение', command=test_connection).pack(side='left')
    tk.Button(btns, text='Отмена', command=cancel_selection).pack(side='right')
    tk.Button(btns, text=action_text, command=run_with_selected).pack(side='right', padx=(0, 8))

    root.protocol('WM_DELETE_WINDOW', cancel_selection)
    root.mainloop()
    return selected['value']


def main():
    global enter_count, exit_count, tracked_objects, next_id
    global LINE_ORIENTATION, LINE_Y, LINE_X, LINE_START_X, LINE_END_X, LINE_START_Y, LINE_END_Y, draw_state
    global STOP_NAME, ROUTE_NAME, VEHICLE_NAME, DOOR_NUMBER
    global STOP_LIST, current_stop_index, frame_height, frame_width, VIDEO_SOURCE, settings_request
    global YOLO_CONFIDENCE, PASSENGER_ID_THRESHOLD, YOLO_EVERY_N_FRAMES, YOLO_IMGSZ, TIME_AT_STOP_SEC

    if pd is None:
        print('Pandas не установлен. Установите pandas: pip install pandas')
    if YOLO is None:
        print('Ошибка: ultralytics не установлен. pip install ultralytics')
        return
    init_database()

    # День 11: загружаем сохранённую конфигурацию (пороги, линия, остановка)
    load_config()

    # Инициализация идентификатора пассажиров
    p_ident = None
    p_db = None
    if PASSENGER_ID_ENABLED and IDENT_AVAILABLE:
        try:
            p_ident = PassengerIdentifier()
            p_db = PassengerDB(DATABASE_PATH, p_ident, PASSENGER_ID_THRESHOLD)
            print('Идентификация пассажиров: активна')
        except Exception as _e_id:
            print(f'Идентификация пассажиров: ошибка инициализации ({_e_id})')
    track_pids: dict = {}       # obj_id → passenger_id (вид ‘P0001’)
    track_id_frame: dict = {}   # obj_id → frame_idx последней идентификации

    # Инициализация анализатора профиля/силуэта (рост, одежда, цвета)
    profile_analyzer = None
    if PROFILE_ANALYSIS_ENABLED and IDENT_AVAILABLE:
        try:
            profile_analyzer = ProfileAnalyzer(CAMERA_HEIGHT_M, VISIBLE_HEIGHT_M)
            print('Анализ профиля/силуэта: активен')
        except Exception as _e_pa:
            print(f'Анализ профиля: ошибка ({_e_pa})')

    # Множество ID пассажиров, находящихся сейчас в транспорте (для траекторий)
    in_vehicle_pids: set = set()
    # Счётчик кадров для обновления параметров внешности
    appearance_frame: dict = {}   # obj_id -> frame_idx последнего анализа внешности
    # Предыдущая остановка — для отслеживания смены и обновления траекторий
    prev_stop_name = STOP_NAME

    # День 15: самодиагностика перед запуском — проверяем все компоненты
    run_self_diagnostic()

    if SHOW_SPLASH_SCREEN:
        show_splash()

    # День 12: инициализируем профилировщик производительности
    profiler = PerformanceProfiler(window_size=120)

    # Загрузка модели YOLO (скачается при первом запуске)
    print(f'Загрузка YOLO модели: {YOLO_MODEL}')
    yolo_model = YOLO(YOLO_MODEL)

    # Выбор источника
    source = choose_source_gui(VIDEO_SOURCE)
    sync_source_settings_from_source(source)
    print(f'Подключение к: {source}')
    cap = open_video_source(source)

    if not cap.isOpened():
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        path = filedialog.askopenfilename(
            title='Видео не найдено — выберите файл',
            filetypes=[('Video', '*.mp4 *.avi *.mov *.mkv'), ('All', '*.*')]
        )
        if path:
            source = path
            sync_source_settings_from_source(source)
            cap = open_video_source(path)
        if not cap.isOpened():
            print('Камера/видео не открыты. Попытка webcam (0)...')
            source = 0
            sync_source_settings_from_source(source)
            cap = open_video_source(0)
        if not cap.isOpened():
            print('Нет источника видео.')
            return

    # ── Оптимизация захвата (с безопасным режимом для USB/web-камер) ───────
    apply_capture_profile(cap, source)

    real_fps = cap.get(cv2.CAP_PROP_FPS)
    print(f'FPS камеры: {real_fps}')

    # Базовый размер для обработки/показа, затем авто-подгонка под экран.
    src_w = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH) or 0)
    src_h = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT) or 0)
    base_w = src_w if src_w > 0 else DISPLAY_W
    base_h = src_h if src_h > 0 else DISPLAY_H
    base_w = min(base_w, DISPLAY_W)
    base_h = min(base_h, DISPLAY_H)
    disp_w, disp_h = fit_frame_to_screen(base_w, base_h, PANEL_H)
    print(f'Размер обработки/показа: {disp_w}x{disp_h} (источник {src_w}x{src_h})')

    # Запускаем веб-панель мониторинга
    if ENABLE_WEB_DASHBOARD:
        start_dashboard()

    # Оборачиваем в многопоточный ридер (фоновое чтение кадров)
    reader = ThreadedVideoReader(cap)
    fps_timer = time.time()
    fps_count = 0
    display_fps = 0
    no_frame_since = None
    last_reconnect_ts = 0.0
    frame_idx = 0

    # День 8: буферизация YOLO — хранит боксы между запусками
    cached_boxes = []             # сохранённые боксы с последнего YOLO-прохода
    yolo_latency_ms = 0.0         # замер времени вывода YOLO (мс)
    loop_latency_ms = 0.0         # замер полного цикла (мс)
    loop_t0 = time.time()         # таймер начала итерации

    WIN = f'People Counter  |  {ROUTE_NAME}  |  {VEHICLE_NAME}'
    cv2.namedWindow(WIN, cv2.WINDOW_AUTOSIZE)
    cv2.setMouseCallback(WIN, mouse_callback)

    # День 11: запускаем сторожевой поток
    watchdog = WatchdogThread(timeout_sec=15.0)
    watchdog.start()

    def refresh_window_title():
        title = f'People Counter  |  {ROUTE_NAME}  |  {VEHICLE_NAME}'
        if hasattr(cv2, 'setWindowTitle'):
            try:
                cv2.setWindowTitle(WIN, title)
            except Exception:
                pass

    def open_settings_during_run():
        nonlocal source, cap, reader, disp_w, disp_h, src_w, src_h, cached_boxes
        nonlocal no_frame_since, last_reconnect_ts, fps_timer, fps_count, display_fps
        global settings_request, tracked_objects, next_id
        global LINE_Y, LINE_X, LINE_START_X, LINE_END_X, LINE_START_Y, LINE_END_Y

        settings_request = False
        previous_source = source
        new_source = choose_source_gui(
            previous_source,
            allow_autostart=False,
            window_title='Настройки программы',
            action_text='Применить',
        )

        if new_source != previous_source:
            print(f'Переключение источника на: {new_source}')
            new_cap = open_video_source(new_source)
            if not new_cap.isOpened():
                print('Новый источник не открылся, продолжаем со старым.')
                try:
                    new_cap.release()
                except Exception:
                    pass
                sync_source_settings_from_source(previous_source)
                save_config()
                refresh_window_title()
                return

            apply_capture_profile(new_cap, new_source)
            new_reader = ThreadedVideoReader(new_cap)
            old_reader = reader
            reader = new_reader
            cap = new_cap
            source = new_source
            old_reader.release()

            cached_boxes = []
            tracked_objects.clear()
            track_pids.clear()
            track_id_frame.clear()
            appearance_frame.clear()
            in_vehicle_pids.clear()
            next_id = 0
            no_frame_since = None
            last_reconnect_ts = 0.0
            fps_timer = time.time()
            fps_count = 0
            display_fps = 0

            src_w = int(new_cap.get(cv2.CAP_PROP_FRAME_WIDTH) or 0)
            src_h = int(new_cap.get(cv2.CAP_PROP_FRAME_HEIGHT) or 0)
            base_w = src_w if src_w > 0 else DISPLAY_W
            base_h = src_h if src_h > 0 else DISPLAY_H
            base_w = min(base_w, DISPLAY_W)
            base_h = min(base_h, DISPLAY_H)
            disp_w, disp_h = fit_frame_to_screen(base_w, base_h, PANEL_H)

            LINE_Y = min(max(LINE_Y, 0), max(0, disp_h - 1))
            LINE_X = min(max(LINE_X, 0), max(0, disp_w - 1))
            LINE_START_X = min(max(LINE_START_X, 0), max(1, disp_w - 1))
            LINE_END_X = min(max(LINE_END_X, 0), max(1, disp_w - 1))
            LINE_START_Y = min(max(LINE_START_Y, 0), max(1, disp_h - 1))
            LINE_END_Y = min(max(LINE_END_Y, 0), max(1, disp_h - 1))
            print(f'Новый источник подключён: {source}')
        else:
            print('Настройки обновлены без смены источника.')

        refresh_window_title()

    while True:
        # День 11: проверяем запрос на завершение (SIGINT/SIGTERM)
        if _shutdown_requested:
            print('[Shutdown] Штатное завершение по сигналу...')
            break
        # День 11: heartbeat для watchdog
        watchdog.ping()
        ret, frame = reader.read()
        if not ret or frame is None:
            # Конец файла — перемотка только для локальных видеофайлов.
            if isinstance(source, str) and os.path.isfile(source):
                reader.cap.set(cv2.CAP_PROP_POS_FRAMES, 0)
            else:
                now_t = time.time()
                if no_frame_since is None:
                    no_frame_since = now_t

                is_network_source = isinstance(source, str) and source.lower().startswith(('rtsp://', 'http://', 'https://'))
                should_reconnect = (
                    AUTO_RECONNECT_STREAM and is_network_source and
                    (now_t - no_frame_since) >= RECONNECT_NOFRAME_SEC and
                    (now_t - last_reconnect_ts) >= RECONNECT_COOLDOWN_SEC
                )

                if should_reconnect:
                    last_reconnect_ts = now_t
                    print('Поток завис/оборвался. Переподключение...')
                    new_cap = open_video_source(source)
                    if new_cap.isOpened():
                        apply_capture_profile(new_cap, source)
                        new_reader = ThreadedVideoReader(new_cap)
                        old_reader = reader
                        reader = new_reader
                        old_reader.release()
                        no_frame_since = None
                        print('Переподключение успешно.')
                    else:
                        try:
                            new_cap.release()
                        except Exception:
                            pass
                        print('Переподключение не удалось, повтор через паузу.')

                time.sleep(0.01)
            continue
        else:
            no_frame_since = None

        frame = apply_camera_rotation(frame)

        # Счётчик FPS
        fps_count += 1
        now_t = time.time()
        if now_t - fps_timer >= 1.0:
            display_fps = fps_count
            fps_count = 0
            fps_timer = now_t

        # Нормализуем разрешение до размера, гарантированно вмещающегося в экран.
        if frame.shape[1] != disp_w or frame.shape[0] != disp_h:
            frame = cv2.resize(frame, (disp_w, disp_h), interpolation=cv2.INTER_AREA)
        H_fr, W_fr = disp_h, disp_w
        frame_height = H_fr  # для mouse_callback
        frame_width = W_fr

        # День 8: замер полного цикла
        loop_t0 = time.time()

        # ── YOLO детекция людей ─────────────────────────────────────────────
        # День 8: пропуск кадров — YOLO запускается каждые YOLO_EVERY_N_FRAMES кадров.
        # В промежуточных кадрах используем сохранённые боксы (cached_boxes).
        run_yolo = (frame_idx % YOLO_EVERY_N_FRAMES == 0)

        if run_yolo:
            _yolo_t0 = time.time()
            results = yolo_model(frame, conf=YOLO_CONFIDENCE, classes=[0],
                                 verbose=False, imgsz=YOLO_IMGSZ, half=YOLO_HALF)
            yolo_latency_ms = (time.time() - _yolo_t0) * 1000.0

            boxes = []
            for r in results:
                for box in r.boxes:
                    x1, y1, x2, y2 = box.xyxy[0].cpu().numpy().astype(int)
                    w, h = x2 - x1, y2 - y1
                    if w * h < MIN_AREA:
                        continue
                    boxes.append((x1, y1, w, h))
            cached_boxes = boxes  # запоминаем для промежуточных кадров
        else:
            boxes = cached_boxes  # используем кешированные боксы

        # Отрисовка боксов
        for (bx, by, bw, bh) in boxes:
            conf_color = (0, 255, 0) if run_yolo else (0, 200, 180)
            cv2.rectangle(frame, (bx, by), (bx + bw, by + bh), conf_color, 2)

        current_centers = []
        for x, y, w, h in boxes:
            cx, cy = get_count_anchor_point(x, y, w, h)
            cv2.circle(frame, (cx, cy), 5, (0, 0, 255), -1)
            current_centers.append((cx, cy))

        # ── Трекинг + подсчёт пересечений ────────────────────────────────────
        # Грeedily сопоставляем текущие центры с существующими треками.
        # Каждый трек = (center, crossed, direction, lost_frames).
        # «lost_frames» позволяет трекам пережить TRACKING_LOST_FRAMES кадров
        # без детекции — это защита от двойного счёта при однокадровом пропуске YOLO.
        max_track_dist = W_fr * TRACKING_MAX_DIST_RATIO
        updated_tracked = {}
        assignments = {}   # индекс центра → obj_id
        used_ids    = set()

        for i, (cx, cy) in enumerate(current_centers):
            best_id, best_d = None, max_track_dist
            for obj_id, td in tracked_objects.items():
                if obj_id in used_ids:
                    continue
                d = np.hypot(cx - td[0][0], cy - td[0][1])
                if d < best_d:
                    best_d, best_id = d, obj_id
            if best_id is not None:
                assignments[i] = best_id
                used_ids.add(best_id)

        for i, (cx, cy) in enumerate(current_centers):
            center = (cx, cy)
            if i in assignments:
                oid = assignments[i]
                prev_cx, prev_cy = tracked_objects[oid][0]
                crossed   = tracked_objects[oid][1]
                direction = tracked_objects[oid][2]

                line_distance = abs(cy - LINE_Y) if LINE_ORIENTATION == 'horizontal' else abs(cx - LINE_X)
                if crossed and line_distance >= LINE_REARM_DISTANCE:
                    crossed = False

                if not crossed:
                    if LINE_ORIENTATION == 'horizontal':
                        if prev_cy < LINE_Y <= cy:
                            count_event = resolve_count_event('in')
                            crossed = True; direction = 'down'
                            if count_event == 'in':
                                enter_count += 1
                            else:
                                exit_count += 1
                            log_event(count_event, 'entry' if count_event == 'in' else 'exit')
                            # Траектория: начинаем путь пассажира при входе
                            if p_db is not None:
                                _pid = track_pids.get(oid, '')
                                if _pid:
                                    if count_event == 'in':
                                        p_db.start_trajectory(_pid, STOP_NAME, ROUTE_NAME)
                                        in_vehicle_pids.add(_pid)
                                        if od_tracker is not None:
                                            od_tracker.log_entry(_pid, STOP_NAME)
                                    else:
                                        p_db.update_exit_stop(_pid, STOP_NAME)
                                        p_db.finish_trajectory(_pid, STOP_NAME)
                                        in_vehicle_pids.discard(_pid)
                                        if od_tracker is not None:
                                            od_tracker.log_exit(_pid, STOP_NAME)
                        elif prev_cy > LINE_Y >= cy:
                            count_event = resolve_count_event('out')
                            crossed = True; direction = 'up'
                            if count_event == 'in':
                                enter_count += 1
                            else:
                                exit_count += 1
                            log_event(count_event, 'entry' if count_event == 'in' else 'exit')
                            # Траектория: завершаем путь пассажира при выходе
                            if p_db is not None:
                                _pid = track_pids.get(oid, '')
                                if _pid:
                                    if count_event == 'in':
                                        p_db.start_trajectory(_pid, STOP_NAME, ROUTE_NAME)
                                        in_vehicle_pids.add(_pid)
                                        if od_tracker is not None:
                                            od_tracker.log_entry(_pid, STOP_NAME)
                                    else:
                                        p_db.update_exit_stop(_pid, STOP_NAME)
                                        p_db.finish_trajectory(_pid, STOP_NAME)
                                        in_vehicle_pids.discard(_pid)
                                        if od_tracker is not None:
                                            od_tracker.log_exit(_pid, STOP_NAME)
                    else:
                        if prev_cx < LINE_X <= cx:
                            count_event = resolve_count_event('in')
                            crossed = True; direction = 'right'
                            if count_event == 'in':
                                enter_count += 1
                            else:
                                exit_count += 1
                            log_event(count_event, 'entry' if count_event == 'in' else 'exit')
                            # Траектория: начинаем путь пассажира при входе
                            if p_db is not None:
                                _pid = track_pids.get(oid, '')
                                if _pid:
                                    if count_event == 'in':
                                        p_db.start_trajectory(_pid, STOP_NAME, ROUTE_NAME)
                                        in_vehicle_pids.add(_pid)
                                        if od_tracker is not None:
                                            od_tracker.log_entry(_pid, STOP_NAME)
                                    else:
                                        p_db.update_exit_stop(_pid, STOP_NAME)
                                        p_db.finish_trajectory(_pid, STOP_NAME)
                                        in_vehicle_pids.discard(_pid)
                                        if od_tracker is not None:
                                            od_tracker.log_exit(_pid, STOP_NAME)
                        elif prev_cx > LINE_X >= cx:
                            count_event = resolve_count_event('out')
                            crossed = True; direction = 'left'
                            if count_event == 'in':
                                enter_count += 1
                            else:
                                exit_count += 1
                            log_event(count_event, 'entry' if count_event == 'in' else 'exit')
                            # Траектория: завершаем путь пассажира при выходе
                            if p_db is not None:
                                _pid = track_pids.get(oid, '')
                                if _pid:
                                    if count_event == 'in':
                                        p_db.start_trajectory(_pid, STOP_NAME, ROUTE_NAME)
                                        in_vehicle_pids.add(_pid)
                                        if od_tracker is not None:
                                            od_tracker.log_entry(_pid, STOP_NAME)
                                    else:
                                        p_db.update_exit_stop(_pid, STOP_NAME)
                                        p_db.finish_trajectory(_pid, STOP_NAME)
                                        in_vehicle_pids.discard(_pid)
                                        if od_tracker is not None:
                                            od_tracker.log_exit(_pid, STOP_NAME)

                updated_tracked[oid] = (center, crossed, direction, 0)
            else:
                updated_tracked[next_id] = (center, False, None, 0)
                next_id += 1

        # Сохраняем потерянные треки ещё TRACKING_LOST_FRAMES кадров
        for oid, td in tracked_objects.items():
            if oid not in used_ids:
                lost = (td[3] if len(td) > 3 else 0) + 1
                if lost <= TRACKING_LOST_FRAMES:
                    updated_tracked[oid] = (td[0], td[1], td[2], lost)

        tracked_objects = updated_tracked

        # ── Идентификация пассажиров ──────────────────────────────────────
        frame_idx += 1
        if p_ident is not None and p_db is not None:
            for i, (cx, cy) in enumerate(current_centers):
                oid = assignments.get(i)
                if oid is None:
                    for k, td in tracked_objects.items():
                        if td[0] == (cx, cy):
                            oid = k
                            break
                if oid is None:
                    continue
                is_new_track = oid not in track_pids
                frames_since = frame_idx - track_id_frame.get(oid, 0)
                if is_new_track or frames_since >= IDENTIFY_EVERY_N_FRAMES:
                    bx, by, bw, bh = boxes[i]
                    # День 7: мультимодальный матчинг (лицо + внешность)
                    face_params = p_ident.extract_face_params(
                        frame, bbox_xyxy=(bx, by, bx + bw, by + bh))
                    appear_desc = None
                    if profile_analyzer is not None:
                        appear_desc = profile_analyzer.compute_appearance_descriptor(
                            frame, (bx, by, bw, bh), H_fr)
                    pid, _ = p_db.match_or_create(
                        face_params, from_stop=STOP_NAME, appear_desc=appear_desc)
                    track_pids[oid] = pid
                    track_id_frame[oid] = frame_idx
            for oid in list(track_pids):
                if oid not in tracked_objects:
                    del track_pids[oid]
            for oid in list(track_id_frame):
                if oid not in tracked_objects:
                    del track_id_frame[oid]

        # ── Отслеживание смены остановки (обновление траекторий) ─────────
        # Когда водитель переключает остановку, добавляем её в траектории
        # всех пассажиров, находящихся сейчас в транспорте.
        if STOP_NAME != prev_stop_name:
            if p_db is not None:
                for _vpid in in_vehicle_pids:
                    p_db.add_stop_to_trajectory(_vpid, STOP_NAME)
            prev_stop_name = STOP_NAME

        # ── Анализ внешности (профиль/силуэт): рост, одежда, цвета ──────
        # Периодически (раз в APPEARANCE_EVERY_N_FRAMES) извлекаем параметры
        # внешности каждого отслеживаемого пассажира и сохраняем в БД.
        if profile_analyzer is not None and p_db is not None:
            for i, (bx, by, bw, bh) in enumerate(boxes):
                oid = assignments.get(i)
                if oid is None:
                    continue
                pid = track_pids.get(oid, '')
                if not pid:
                    continue
                # Не анализируем слишком часто — только каждые N кадров
                frames_since_app = frame_idx - appearance_frame.get(oid, 0)
                if frames_since_app >= APPEARANCE_EVERY_N_FRAMES:
                    app_params = profile_analyzer.extract_full_appearance(
                        frame, (bx, by, bw, bh), H_fr)
                    if app_params is not None:
                        p_db.save_appearance(pid, app_params)
                        appearance_frame[oid] = frame_idx
            # Очистка appearance_frame для удалённых треков
            for oid in list(appearance_frame):
                if oid not in tracked_objects:
                    del appearance_frame[oid]

        # ── Наложение ID пассажиров на бокс ─────────────────────────────
        for i, (bx, by, bw, bh) in enumerate(boxes):
            oid = assignments.get(i)
            if oid is not None:
                pid = track_pids.get(oid, '')
                if pid:
                    cv2.rectangle(frame, (bx, by), (bx + bw, by + bh), (0, 180, 255), 2)
                    cv2.putText(frame, pid, (bx, by - 22),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 200, 255), 2, cv2.LINE_AA)

        # Рисование линии
        if LINE_ORIENTATION == 'horizontal':
            cv2.line(frame, (LINE_START_X, LINE_Y), (LINE_END_X, LINE_Y), (255, 0, 0), 2)
        else:
            lx = min(max(LINE_X, 0), W_fr - 1)
            cv2.line(frame, (lx, LINE_START_Y), (lx, LINE_END_Y), (255, 0, 0), 2)

        if draw_state['drawing'] and draw_state['start'] and draw_state['end']:
            cv2.line(frame, draw_state['start'], draw_state['end'], (0, 255, 255), 1)

        # ── Информационная панель (редизайн) ───────────────────────────────────
        in_salon = max(0, enter_count - exit_count)
        panel_h  = PANEL_H
        panel = get_panel_background(W_fr, panel_h).copy()
        panel_texts = []

        def pp(text, x, y, size=0.55, color=(220, 220, 220), bold=False):
            font_px = max(14, int(34 * size) + (4 if bold else 0))
            panel_texts.append({
                'text': text,
                'x': x,
                'y': y - font_px + 8,
                'font_px': font_px,
                'color': color,
                'center': False,
            })

        pp(f'{ORG_NAME}  |  {ROUTE_NAME}  |  {VEHICLE_NAME}  |  Дверь {DOOR_NUMBER}',
           14, 30, 0.62, (75, 220, 255), bold=True)
        pp(f'Остановка [{current_stop_index+1}/{len(STOP_LIST)}]: {STOP_NAME}', 14, 60, 0.56, (235, 235, 235))

        # Кнопки переключения остановок (мышь)
        btn_h, btn_w = 24, 36
        btn_y = 44
        # кнопка ◀ предыдущая
        bpx = W_fr - 150
        _c_prev = (60, 160, 60) if current_stop_index > 0 else (50, 50, 50)
        cv2.rectangle(panel, (bpx, btn_y), (bpx + btn_w, btn_y + btn_h), _c_prev, -1)
        cv2.rectangle(panel, (bpx, btn_y), (bpx + btn_w, btn_y + btn_h), (100, 200, 100), 1)
        cv2.putText(panel, '<', (bpx + 12, btn_y + 18), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (220, 255, 220), 2, cv2.LINE_AA)
        panel_buttons['prev'] = (bpx, btn_y, bpx + btn_w, btn_y + btn_h)
        # кнопка ▶ следующая
        bnx = bpx + btn_w + 6
        _c_next = (60, 160, 60) if current_stop_index < len(STOP_LIST) - 1 else (50, 50, 50)
        cv2.rectangle(panel, (bnx, btn_y), (bnx + btn_w, btn_y + btn_h), _c_next, -1)
        cv2.rectangle(panel, (bnx, btn_y), (bnx + btn_w, btn_y + btn_h), (100, 200, 100), 1)
        cv2.putText(panel, '>', (bnx + 12, btn_y + 18), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (220, 255, 220), 2, cv2.LINE_AA)
        panel_buttons['next'] = (bnx, btn_y, bnx + btn_w, btn_y + btn_h)
        # кнопка […] управление
        bmx = bnx + btn_w + 6
        cv2.rectangle(panel, (bmx, btn_y), (bmx + btn_w, btn_y + btn_h), (80, 60, 30), -1)
        cv2.rectangle(panel, (bmx, btn_y), (bmx + btn_w, btn_y + btn_h), (180, 150, 60), 1)
        cv2.putText(panel, '...', (bmx + 4, btn_y + 18), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 230, 150), 1, cv2.LINE_AA)
        panel_buttons['manage'] = (bmx, btn_y, bmx + btn_w, btn_y + btn_h)
        # кнопка SET настройки
        bsx = bmx + btn_w + 6
        btn_w_settings = 52
        cv2.rectangle(panel, (bsx, btn_y), (bsx + btn_w_settings, btn_y + btn_h), (40, 72, 110), -1)
        cv2.rectangle(panel, (bsx, btn_y), (bsx + btn_w_settings, btn_y + btn_h), (120, 190, 255), 1)
        cv2.putText(panel, 'SET', (bsx + 8, btn_y + 18), cv2.FONT_HERSHEY_SIMPLEX, 0.48, (220, 245, 255), 1, cv2.LINE_AA)
        panel_buttons['settings'] = (bsx, btn_y, bsx + btn_w_settings, btn_y + btn_h)

        # Карточки счетчиков
        card_y1, card_y2 = 76, 160
        card_w, gap = 180, 10
        x1 = 14
        x2 = x1 + card_w + gap
        x3 = x2 + card_w + gap
        x4 = x3 + card_w + gap
        _uniq_pax = p_db.get_unique_count() if p_db is not None else 0

        cv2.rectangle(panel, (x1, card_y1), (x1 + card_w, card_y2), (36, 80, 42), -1)
        cv2.rectangle(panel, (x2, card_y1), (x2 + card_w, card_y2), (62, 56, 24), -1)
        cv2.rectangle(panel, (x3, card_y1), (x3 + card_w, card_y2), (65, 44, 44), -1)
        cv2.rectangle(panel, (x4, card_y1), (x4 + card_w, card_y2), (28, 56, 80), -1)

        cv2.rectangle(panel, (x1, card_y1), (x1 + card_w, card_y2), (70, 170, 100), 1)
        cv2.rectangle(panel, (x2, card_y1), (x2 + card_w, card_y2), (180, 150, 60), 1)
        cv2.rectangle(panel, (x3, card_y1), (x3 + card_w, card_y2), (190, 120, 120), 1)
        cv2.rectangle(panel, (x4, card_y1), (x4 + card_w, card_y2), (80, 160, 220), 1)

        pp('Вошло', x1 + 10, 100, 0.55, (160, 235, 190), bold=True)
        pp(str(enter_count), x1 + 12, 143, 1.0, (200, 255, 220), bold=True)

        pp('Вышло', x2 + 10, 100, 0.55, (255, 230, 150), bold=True)
        pp(str(exit_count), x2 + 12, 143, 1.0, (255, 240, 190), bold=True)

        pp('В салоне', x3 + 10, 100, 0.55, (255, 180, 180), bold=True)
        pp(str(in_salon), x3 + 12, 143, 1.0, (255, 215, 215), bold=True)

        pp('Уник. пасс.', x4 + 10, 100, 0.50, (140, 200, 255), bold=True)
        pp(str(_uniq_pax), x4 + 12, 143, 1.0, (160, 220, 255), bold=True)

        ts_now = datetime.now().strftime('%H:%M:%S  %d.%m.%Y')
        pp(ts_now, W_fr - 235, 28, 0.54, (170, 190, 210))
        pp(f'FPS: {display_fps}', W_fr - 235, 56, 0.50, (100, 200, 100))
        # День 8: показываем latency YOLO и цикла
        pp(f'YOLO: {yolo_latency_ms:.0f}ms  Loop: {loop_latency_ms:.0f}ms',
           W_fr - 235, 80, 0.42, (160, 160, 100))
        # День 9: показываем текущие пороги калибровки
        pp(f'YOLOconf={YOLO_CONFIDENCE:.2f}  IDthr={PASSENGER_ID_THRESHOLD:.2f}',
           W_fr - 235, 100, 0.40, (140, 140, 160))
        # Пожелание заказчика: показываем порог времени стоянки для транзитов
        pp(f'TimeAtStop={TIME_AT_STOP_SEC:.0f}сек',
           W_fr - 235, 120, 0.39, (160, 140, 140))
        pp('PgUp/PgDn=остановки  N=настр.  Стрелки/WASD=линия  G=список  R=отчёт  Q=выход',
           14, panel_h - 14, 0.46, (160, 160, 165))
        draw_text_unicode_batch(panel, panel_texts)

        # Обновляем данные для веб-панели
        total_in, total_out = get_total_counters()
        
        # OD-матрица HTML
        od_html = ''
        if ENABLE_WEB_DASHBOARD and od_tracker is not None:
            try:
                od_html = od_tracker.get_od_html_table()
            except:
                pass
        
        # Таблица параметров пассажиров
        pax_html = ''
        if ENABLE_WEB_DASHBOARD and p_db is not None:
            try:
                pax_html = p_db.get_passengers_html_table()
            except:
                pass
        
        live_data.update({
            'enter': enter_count,
            'exit': exit_count,
            'in_salon': in_salon,
            'total_enter': total_in,
            'total_exit': total_out,
            'unique_passengers': (p_db.get_unique_count() if p_db is not None else 0),
            'stop': STOP_NAME,
            'route': ROUTE_NAME,
            'vehicle': VEHICLE_NAME,
            'door': DOOR_NUMBER,
            'time': ts_now,
            'fps': display_fps,
            'stop_index': current_stop_index + 1,
            'stop_total': len(STOP_LIST),
            'stops': STOP_LIST[:],
            'stop_stats': get_stop_stats(),
            'od_html': od_html,
            'pax_html': pax_html,
        })

        # День 8: замер полного цикла
        loop_latency_ms = (time.time() - loop_t0) * 1000.0

        # День 12: записываем замеры в профилировщик
        profiler.record_loop(loop_latency_ms / 1000.0, len(tracked_objects))

        # День 12: периодическая очистка устаревших треков (экономия ОЗУ)
        if frame_idx > 0 and frame_idx % TRACK_CLEANUP_FRAMES == 0:
            cleanup_stale_tracks(
                tracked_objects, track_pids, track_id_frame,
                appearance_frame, frame_idx
            )

        combined = np.vstack([frame, panel])
        cv2.imshow(WIN, combined)

        key = cv2.waitKeyEx(1)
        if key_is(key, 'q', 'Q', 'й', 'Й'):
            break
        elif key_is(key, 'h', 'H', 'р', 'Р'):
            LINE_ORIENTATION = 'horizontal'
        elif key_is(key, 'v', 'V', 'м', 'М'):
            LINE_ORIENTATION = 'vertical'
        elif key == 2490368 or key_is(key, 'w', 'W', 'ц', 'Ц'):
            if LINE_ORIENTATION == 'horizontal':
                LINE_Y = max(0, LINE_Y - LINE_KEYBOARD_STEP)
            else:
                LINE_X = max(0, LINE_X - LINE_KEYBOARD_STEP)
        elif key == 2621440 or key_is(key, 's', 'S', 'ы', 'Ы'):
            if LINE_ORIENTATION == 'horizontal':
                LINE_Y = min(H_fr - 1, LINE_Y + LINE_KEYBOARD_STEP)
            else:
                LINE_X = min(W_fr - 1, LINE_X + LINE_KEYBOARD_STEP)
        elif key == 2424832 or key_is(key, 'a', 'A', 'ф', 'Ф'):
            if LINE_ORIENTATION == 'vertical':
                LINE_X = max(0, LINE_X - LINE_KEYBOARD_STEP)
            else:
                LINE_Y = max(0, LINE_Y - LINE_KEYBOARD_STEP)
        elif key == 2555904 or key_is(key, 'd', 'D', 'в', 'В'):
            if LINE_ORIENTATION == 'vertical':
                LINE_X = min(W_fr - 1, LINE_X + LINE_KEYBOARD_STEP)
            else:
                LINE_Y = min(H_fr - 1, LINE_Y + LINE_KEYBOARD_STEP)
        elif key_is(key, 'c', 'C', 'с', 'С'):
            LINE_Y = H_fr // 2
            LINE_X = W_fr // 2
            LINE_START_X, LINE_END_X = 0, W_fr
            LINE_START_Y, LINE_END_Y = 0, H_fr
            draw_state.update({'drawing': False, 'start': None, 'end': None})
            enter_count = 0
            exit_count  = 0
            tracked_objects.clear()
        elif key == 2162688 or key_is(key, ',', '<', 'б', 'Б'):  # PgUp
            switch_to_stop(current_stop_index - 1)
        elif key == 2228224 or key_is(key, '.', '>', 'ю', 'Ю'):  # PgDn
            switch_to_stop(current_stop_index + 1)
        elif key_is(key, 'g', 'G', 'п', 'П'):  # G/П — окно управления остановками
            manage_stops_gui()
        elif key_is(key, 'n', 'N', 'т', 'Т'):
            settings_request = True
        elif key_is(key, 'r', 'R', 'к', 'К'):
            if pd is not None:
                generate_reports()
            else:
                print('pandas не установлен')
            # OD-матрица: экспортируем матрицу
            if od_tracker is not None:
                try:
                    od_tracker.export_od_json(app_path('od_matrix.json'))
                    od_tracker.export_od_csv(app_path('od_matrix.csv'))
                    print(f'OD-матрица экспортирована: {app_path("od_matrix.json")}, {app_path("od_matrix.csv")}')
                except Exception as ex:
                    print(f'Ошибка экспорта OD: {ex}')
        elif key_is(key, 'x', 'X', 'ч', 'Ч'):  # X — ночная сшивка маршрутов
            print('Запуск сшивки траекторий через маршруты...')
            try:
                from cross_route_stitcher import main as stitch_main
                stitch_main()
            except Exception as ex:
                print(f'Ошибка сшивки: {ex}')
        # ── День 12: профиль производительности по клавише P ──────────
        elif key_is(key, 'p', 'P', 'з', 'З'):
            num_cached = len(p_db._cache) if p_db is not None and hasattr(p_db, '_cache') else 0
            report = profiler.format_report(len(tracked_objects), num_cached)
            print(report)
        # ── День 13: создание файлов сборки EXE по клавише B ─────────
        elif key_is(key, 'b', 'B', 'и', 'И'):
            generate_build_script()
        # ── День 14: генерация инструкции пользователя по клавише M ──
        elif key_is(key, 'm', 'M', 'ь', 'Ь'):
            generate_user_manual()
        # ── День 9: калибровка порогов горячими клавишами ──────────────
        elif key == ord('[') or key == ord('х'):  # [ — YOLO_CONFIDENCE −0.05
            YOLO_CONFIDENCE = max(0.10, round(YOLO_CONFIDENCE - 0.05, 2))
            print(f'YOLO_CONFIDENCE = {YOLO_CONFIDENCE}')
        elif key == ord(']') or key == ord('ъ'):  # ] — YOLO_CONFIDENCE +0.05
            YOLO_CONFIDENCE = min(0.95, round(YOLO_CONFIDENCE + 0.05, 2))
            print(f'YOLO_CONFIDENCE = {YOLO_CONFIDENCE}')
        elif key == ord('-'):  # − — ID_THRESHOLD −0.02
            PASSENGER_ID_THRESHOLD = max(0.05, round(PASSENGER_ID_THRESHOLD - 0.02, 2))
            if p_db is not None:
                p_db._threshold = PASSENGER_ID_THRESHOLD
            print(f'PASSENGER_ID_THRESHOLD = {PASSENGER_ID_THRESHOLD}')
        elif key == ord('='):  # = — ID_THRESHOLD +0.02
            PASSENGER_ID_THRESHOLD = min(0.60, round(PASSENGER_ID_THRESHOLD + 0.02, 2))
            if p_db is not None:
                p_db._threshold = PASSENGER_ID_THRESHOLD
            print(f'PASSENGER_ID_THRESHOLD = {PASSENGER_ID_THRESHOLD}')
        # ── Настройка порога времени стоянки для выявления транзитов ──────────
        elif key == ord(';') or key == ord(':'):  # ; — TIME_AT_STOP_SEC −5 сек
            TIME_AT_STOP_SEC = max(5.0, TIME_AT_STOP_SEC - 5.0)
            print(f'TIME_AT_STOP_SEC = {TIME_AT_STOP_SEC}сек (стоянка для выявления транзитов)')
        elif key == ord('\'') or key == ord('"'):  # ' — TIME_AT_STOP_SEC +5 сек
            TIME_AT_STOP_SEC = min(300.0, TIME_AT_STOP_SEC + 5.0)
            print(f'TIME_AT_STOP_SEC = {TIME_AT_STOP_SEC}сек (стоянка для выявления транзитов)')

        if settings_request:
            open_settings_during_run()

    # День 11: останавливаем watchdog
    watchdog.stop()

    reader.release()
    cv2.destroyAllWindows()

    # День 11: сохраняем конфигурацию при выходе
    save_current_stop_counters()
    save_config()
    print('Конфигурация сохранена.')

    total_in, total_out = get_total_counters()
    in_salon_final = max(0, total_in - total_out)
    print('\n' + '='*50)
    print(f'  {ORG_NAME} — итоги маршрута')
    print('='*50)
    print(f'  Маршрут:  {ROUTE_NAME}')
    print(f'  ТС:       {VEHICLE_NAME}')
    print(f'  --- По остановкам ---')
    for s in STOP_LIST:
        c = stop_counters.get(s, {'enter': 0, 'exit': 0})
        print(f'    {s}: вошло {c["enter"]}, вышло {c["exit"]}')
    print(f'  --- Итого ---')
    print(f'  Вошло:    {total_in}')
    print(f'  Вышло:    {total_out}')
    print(f'  В салоне: {in_salon_final}')
    print('='*50)

    # День 15: итоговая сводка сессии с рекомендациями
    print_session_summary(profiler, p_db, od_tracker)


if __name__ == "__main__":
    main()
    if pd is not None:
        print('Формируем отчёты...')
        generate_reports()
    else:
        print('Pandas отсутствует: отчёты не сформированы.')