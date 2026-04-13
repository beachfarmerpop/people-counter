import pandas as pd
import sqlite3
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

DB_PATH = 'passenger_flow.db'
REPORT_FOLDER = 'reports'
XLSX_PATH = os.path.join(REPORT_FOLDER, 'passenger_flow_report.xlsx')

os.makedirs(REPORT_FOLDER, exist_ok=True)

conn = sqlite3.connect(DB_PATH)
df = pd.read_sql_query('SELECT * FROM passenger_flow', conn)
conn.close()

if df.empty:
    print('Нет данных в базе. Запусти people_counter.py и пройди через линию.')
    exit()

df['timestamp'] = pd.to_datetime(df['timestamp'])
df['date'] = df['timestamp'].dt.date
df['hour'] = df['timestamp'].dt.hour
df['month'] = df['timestamp'].dt.month
df['year'] = df['timestamp'].dt.year

ctrl = df.groupby(['route', 'vehicle', 'door', 'direction']).agg(total=('id', 'count')).reset_index()
ctrl.columns = ['Маршрут', 'ТС', 'Дверь', 'Направление', 'Кол-во']

sheets = {
    'Все данные': df[['id', 'timestamp', 'route', 'vehicle', 'stop', 'door', 'direction', 'event_type']].rename(columns={
        'id': 'ID', 'timestamp': 'Дата/Время', 'route': 'Маршрут',
        'vehicle': 'ТС', 'stop': 'Остановка', 'door': 'Дверь',
        'direction': 'Направление', 'event_type': 'Тип события'
    }),
    'По дням': df.groupby('date').size().reset_index(name='Кол-во пассажиров').rename(columns={'date': 'Дата'}),
    'По часам': df.groupby(['date', 'hour']).size().reset_index(name='Кол-во').rename(columns={'date': 'Дата', 'hour': 'Час'}),
    'По месяцам': df.groupby(['year', 'month']).size().reset_index(name='Кол-во').rename(columns={'year': 'Год', 'month': 'Месяц'}),
    'По годам': df.groupby('year').size().reset_index(name='Кол-во пассажиров').rename(columns={'year': 'Год'}),
    'По остановкам': df.groupby('stop').size().reset_index(name='Кол-во').rename(columns={'stop': 'Остановка'}),
    'По маршрутам': df.groupby(['route', 'door']).size().reset_index(name='Кол-во').rename(columns={'route': 'Маршрут', 'door': 'Дверь'}),
    'Панель руководителя': ctrl
}

writer = pd.ExcelWriter(XLSX_PATH, engine='openpyxl')

# Цвета заголовков для каждого листа
header_colors = {
    'Все данные':           '1F4E79',
    'По дням':              '2E75B6',
    'По часам':             '2E75B6',
    'По месяцам':           '375623',
    'По годам':             '375623',
    'По остановкам':        '7B2C2C',
    'По маршрутам':         '4472C4',
    'Панель руководителя':  '833C00',
}

row_fill_odd  = PatternFill('solid', fgColor='DCE6F1')
row_fill_even = PatternFill('solid', fgColor='FFFFFF')
thin = Side(border_style='thin', color='CCCCCC')
cell_border = Border(top=thin, left=thin, right=thin, bottom=thin)

for sheet_name, data in sheets.items():
    data.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
    ws = writer.sheets[sheet_name]

    # Заголовок листа (строка 1)
    title_fill = PatternFill('solid', fgColor=header_colors.get(sheet_name, '1F4E79'))
    title_font = Font(bold=True, size=14, color='FFFFFF')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(data.columns))
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = sheet_name
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Заголовки столбцов (строка 2)
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor=header_colors.get(sheet_name, '1F4E79'))
    for cell in ws[2]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = cell_border
    ws.row_dimensions[2].height = 22

    # Данные (строки начиная с 3)
    for row_idx, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row), start=0):
        fill = row_fill_odd if row_idx % 2 == 0 else row_fill_even
        for cell in row:
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = cell_border

    # Ширина столбцов
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 45)

    ws.freeze_panes = 'A3'

writer.close()
print('='*55)
print('Красиво оформленный Excel-отчёт готов!')
print(f'Файл: {os.path.abspath(XLSX_PATH)}')
print('='*55)
print('Листы:')
for name in sheets:
    print(f'  - {name}')
